"""
EY Contact Centre Transformation Business Case Tool v7
═══════════════════════════════════════════════════════
HYBRID ENGINE: Pool-based netting + stepped realization
- Pools set auditable ceilings per lever
- Stepped realization degrades per-initiative (90%/75%/60%/45%)
- Yearly ramp targets with monthly S-curve interpolation
- Role-level FTE breakdown with reskilling matrix
- Risk register auto-generated from initiative profiles
- Full financials: NPV, IRR, scenarios, sensitivity, CX revenue, cost of inaction
"""
from flask import Flask, request, jsonify, render_template, send_file
import uuid, copy, math, io, json, os
from datetime import datetime

app = Flask(__name__)
app.config['SECRET_KEY'] = 'ey-cc-v7'

_store = {}
def _uid(): return uuid.uuid4().hex[:8]
def _get_store():
    if 'enterprise' not in _store:
        _store['enterprise'] = default_enterprise()
    return _store['enterprise']
def _last_result():
    return _store.get('_last_result')

CHANNELS = ['voice','chat','email','portal']
PROCESSES = ['orders','warranty','billing','parts','technical_support','dispatch','returns']
LOCATION_TYPES = ['onshore','nearshore','offshore']

# ═══════════════════════════════════════════════════════
# ROLE MODEL — for workforce breakdown
# ═══════════════════════════════════════════════════════
ROLE_DEFAULTS = [
    {'role':'Agent L1','pct_of_fte':0.55,'cost_multiplier':1.0,'reducible':True},
    {'role':'Agent L2','pct_of_fte':0.18,'cost_multiplier':1.25,'reducible':True},
    {'role':'Agent L3 / SME','pct_of_fte':0.05,'cost_multiplier':1.50,'reducible':False},
    {'role':'QA Analyst','pct_of_fte':0.04,'cost_multiplier':1.15,'reducible':True},
    {'role':'Supervisor','pct_of_fte':0.08,'cost_multiplier':1.40,'reducible':True},
    {'role':'Back-Office','pct_of_fte':0.06,'cost_multiplier':0.90,'reducible':True},
    {'role':'WFM / Planner','pct_of_fte':0.02,'cost_multiplier':1.30,'reducible':False},
    {'role':'Trainer','pct_of_fte':0.02,'cost_multiplier':1.20,'reducible':False},
]

RESKILLING_PATHS = [
    {'from_role':'Agent L1','to_role':'Agent L2','skills':'Advanced product, escalation handling','weeks':4,'cost':3000},
    {'from_role':'Agent L1','to_role':'QA Analyst','skills':'QA methodology, analytics','weeks':6,'cost':4000},
    {'from_role':'Agent L1','to_role':'Chat Specialist','skills':'Multi-chat handling, digital tone','weeks':2,'cost':1500},
    {'from_role':'Agent L2','to_role':'Knowledge Manager','skills':'Content creation, taxonomy','weeks':4,'cost':3500},
    {'from_role':'Back-Office','to_role':'RPA Manager','skills':'Process mapping, automation tools','weeks':8,'cost':6000},
    {'from_role':'QA Analyst','to_role':'AI Trainer','skills':'ML concepts, labeling, model feedback','weeks':6,'cost':5000},
]

# ═══════════════════════════════════════════════════════
# STEPPED REALIZATION FACTORS
# ═══════════════════════════════════════════════════════
# Each subsequent initiative consuming from the same pool/lever gets less effective
STEPPED_REALIZATION = [0.90, 0.75, 0.60, 0.45, 0.35, 0.30, 0.25, 0.20]

def _realization_factor(initiative_index_in_lever):
    """Return stepped realization for the Nth initiative on a lever."""
    if initiative_index_in_lever < len(STEPPED_REALIZATION):
        return STEPPED_REALIZATION[initiative_index_in_lever]
    return STEPPED_REALIZATION[-1]  # floor at last value

# ═══════════════════════════════════════════════════════
# POOL CEILINGS — derived from queue parameters
# ═══════════════════════════════════════════════════════
POOL_TYPES = ['deflection','repeat_reduction','channel_shift','aht_reduction',
              'acw_reduction','concurrency_uplift','shrinkage_reduction',
              'occupancy_uplift','schedule_efficiency']

def compute_pool_ceilings(queues, enterprise):
    """Compute the maximum addressable opportunity per lever across all queues."""
    pools = {}
    for lever in POOL_TYPES:
        pool_contacts = 0
        pool_fte = 0
        pool_detail = []
        for bu, q in queues:
            bl = calc_queue_baseline(q)
            vol = q['monthly_volume']
            fte = bl['fte']
            ch = q['channel']

            if lever == 'deflection':
                # Eligible: simple/medium complexity contacts on channels that support self-service
                eligible_pct = q.get('deflection_eligible_pct') or _default_deflection_eligible(q)
                containment = q.get('containment_feasibility') or _default_containment(q)
                ceiling = vol * eligible_pct * containment
                pool_contacts += ceiling
                if vol > 0:
                    pool_fte += fte * (ceiling / vol)

            elif lever == 'repeat_reduction':
                repeat_pct = q.get('repeat_contact_pct', 0.15)
                # Can reduce up to 70% of repeats
                ceiling = vol * repeat_pct * 0.70
                pool_contacts += ceiling
                if vol > 0:
                    pool_fte += fte * (ceiling / vol)

            elif lever == 'channel_shift':
                if ch == 'voice':
                    # Can shift up to 30% of voice to digital
                    ceiling = vol * 0.30
                    pool_contacts += ceiling
                    if vol > 0:
                        pool_fte += fte * (ceiling / vol)

            elif lever == 'aht_reduction':
                # AHT floor: can't reduce below 40% of current
                aht = q['handle_time_minutes']
                aht_floor = aht * 0.40
                max_reduction_pct = (aht - aht_floor) / max(aht, 0.1)
                pool_fte += fte * max_reduction_pct
                pool_contacts += vol  # all contacts benefit from AHT reduction

            elif lever == 'acw_reduction':
                acw = q.get('after_call_work_minutes', 0)
                acw_floor = acw * 0.20
                if acw > 0:
                    max_reduction_pct = (acw - acw_floor) / max(acw, 0.1)
                    pool_fte += fte * max_reduction_pct * (acw / max(q['handle_time_minutes'] + acw, 1))

            elif lever == 'concurrency_uplift':
                if ch == 'chat':
                    current = q.get('chat_concurrency', 2.5)
                    ceiling_conc = 6.0
                    if current < ceiling_conc:
                        uplift_pct = (ceiling_conc - current) / max(current, 1)
                        pool_fte += fte * min(uplift_pct / (1 + uplift_pct), 0.50)

            elif lever == 'shrinkage_reduction':
                shrink = q.get('shrinkage_pct', 0.30)
                shrink_floor = 0.12
                if shrink > shrink_floor:
                    reduction = (shrink - shrink_floor) / (1 - shrink_floor)
                    pool_fte += fte * reduction * 0.5  # partial FTE impact

            elif lever == 'occupancy_uplift':
                occ = q.get('occupancy_target', 0.82)
                occ_ceil = 0.92 if ch == 'voice' else (0.90 if ch == 'chat' else 0.95)
                if occ < occ_ceil:
                    pool_fte += fte * (occ_ceil - occ)

            elif lever == 'schedule_efficiency':
                sched = q.get('schedule_efficiency', 0.90)
                sched_ceil = 0.98
                if sched < sched_ceil:
                    pool_fte += fte * (sched_ceil - sched) / max(sched, 0.5)

        pools[lever] = {
            'ceiling_contacts': round(pool_contacts, 0),
            'ceiling_fte': round(pool_fte, 2),
            'consumed_contacts': 0,
            'consumed_fte': 0,
            'remaining_contacts': round(pool_contacts, 0),
            'remaining_fte': round(pool_fte, 2),
            'initiatives_consuming': 0,
        }
    return pools

def _default_deflection_eligible(q):
    """Default deflection eligibility based on channel and complexity."""
    ch = q['channel']
    pc = _get_process_complexity(q)
    if ch in ['email', 'portal']:
        return 0.05  # already digital — low deflection potential
    base = 0.35 if pc <= 0.3 else (0.20 if pc <= 0.6 else 0.08)
    return base

def _default_containment(q):
    """Default containment feasibility."""
    pc = _get_process_complexity(q)
    return 0.80 if pc <= 0.3 else (0.55 if pc <= 0.6 else 0.25)


# ═══════════════════════════════════════════════════════
# ENTERPRISE DEFAULTS
# ═══════════════════════════════════════════════════════
def default_enterprise():
    return {
        'program_name': 'Contact Centre Transformation',
        'objective': 'Model the financial and operational impact of contact centre transformation initiatives.',
        'waterfall_order': ['automation_ai', 'opmodel', 'location'],
        'waterfall_active': {'automation_ai': True, 'opmodel': True, 'location': True},
        'planning_horizon_years': 3,
        'discount_rate': 0.10,
        'global_volume_growth_pct': 0.02,
        'global_wage_inflation_pct': 0.03,
        'attrition_rate_monthly': 0.03,
        'redeployment_pct': 0.10,
        'enterprise_fte': 0,
        'currency': 'USD',
        'custom_channels': [],
        'custom_processes': [],
        'location_mix': {
            'onshore': {'fte': 0, 'pct': 0},
            'nearshore': {'fte': 0, 'pct': 0},
            'offshore': {'fte': 0, 'pct': 0},
            '3rd_party': {'fte': 0, 'pct': 0},
        },
        'salary_rates': {
            'onshore': 60000, 'nearshore': 30000, 'offshore': 18000, '3rd_party': 45000
        },
        'sourcing_mix': {'payroll': {'fte': 0, 'pct': 0}, '3rd_party': {'fte': 0, 'pct': 0}},
        'roles': [dict(r) for r in ROLE_DEFAULTS],
        'cx_revenue': {
            'customer_base': 0, 'revenue_per_customer': 0,
            'annual_churn_rate_pct': 0.12, 'churn_improvement_pct': 0.0,
            'csat_baseline': 3.5, 'csat_target': 4.2,
            'fcr_baseline_pct': 0.70, 'fcr_target_pct': 0.82,
        },
        'implementation_costs': {
            'change_management': 0, 'training': 0,
            'integration': 0, 'contingency_pct': 0.10,
        },
        'scenario_multipliers': {
            'conservative': {'adoption': 0.70, 'volume_growth': 1.25, 'benefit_realization': 0.75},
            'base': {'adoption': 1.0, 'volume_growth': 1.0, 'benefit_realization': 1.0},
            'aggressive': {'adoption': 1.20, 'volume_growth': 0.80, 'benefit_realization': 1.15},
        },
        'business_units': [],
        'initiatives_auto': [],
        'initiatives_opmodel': [],
        'location_strategy': [],
        'technology': [],
        'kpis': [],
    }

def default_bu(name='New Business Unit'):
    return {
        'id': _uid(), 'bu_name': name, 'cost_per_fte': 55000,
        'paid_hours_per_fte': 1920, 'shrinkage_pct': 0.30,
        'occupancy_voice': 0.82, 'occupancy_chat': 0.78,
        'channels': ['voice','chat','email'],
        'processes': ['orders','warranty','billing'],
        'total_monthly_volume': 0, 'current_fte': 0,
        'bu_growth_rate': 0.02, 'bu_attrition_rate': 0.03,
        'volume_matrix': {}, 'queues': [],
    }

def default_queue(channel, process, volume, bu):
    occ = bu.get('occupancy_voice',0.82) if channel=='voice' else (bu.get('occupancy_chat',0.78) if channel=='chat' else 0.85)
    return {
        'id': _uid(), 'channel': channel, 'process_tag': process,
        'queue_name': f"{channel.title()} - {process.replace('_',' ').title()}",
        'monthly_volume': volume,
        'handle_time_minutes': {'voice':7.0,'chat':10.0,'email':15.0,'portal':8.0}.get(channel,7.0),
        'after_call_work_minutes': {'voice':2.0,'chat':1.5,'email':1.0,'portal':0.5}.get(channel,2.0),
        'chat_concurrency': 2.5 if channel=='chat' else 1.0,
        'shrinkage_pct': bu.get('shrinkage_pct',0.30),
        'occupancy_target': occ, 'schedule_efficiency': 0.90,
        'async_productivity_factor': 0.75 if channel in ['email','portal'] else 1.0,
        'sla_target': 0.80, 'sla_threshold_seconds': 20 if channel=='voice' else 30,
        'repeat_contact_pct': 0.15, 'transfer_pct': 0.08, 'fcr_pct': 0.72,
        'csat_score': 3.8, 'abandon_rate': 0.05 if channel=='voice' else 0.03,
        'backlog_volume': 0 if channel in ['voice','chat'] else 200,
        'complexity': {'voice':0.7,'chat':0.5,'email':0.4,'portal':0.3}.get(channel,0.5),
        'deflection_eligible_pct': None,  # auto-derived if None
        'containment_feasibility': None,  # auto-derived if None
        'allowed_locations': ['onshore','nearshore','offshore'],
        'paid_hours_per_fte': bu.get('paid_hours_per_fte',1920),
    }

# ═══════════════════════════════════════════════════════
# KPI LIBRARY (same as v6)
# ═══════════════════════════════════════════════════════
def default_kpi_library():
    return [
        {'name':'AHT','unit':'minutes','channels':['voice'],'processes':['all'],'benchmark':5.0,'impact':'decrease','category':'channel'},
        {'name':'AHT','unit':'minutes','channels':['chat'],'processes':['all'],'benchmark':8.0,'impact':'decrease','category':'channel'},
        {'name':'AHT','unit':'minutes','channels':['email'],'processes':['all'],'benchmark':12.0,'impact':'decrease','category':'channel'},
        {'name':'FCR','unit':'%','channels':['voice','chat'],'processes':['all'],'benchmark':78,'impact':'increase','category':'channel'},
        {'name':'CSAT','unit':'score','channels':['voice','chat','email','portal'],'processes':['all'],'benchmark':4.2,'impact':'increase','category':'channel'},
        {'name':'Transfer Rate','unit':'%','channels':['voice','chat'],'processes':['all'],'benchmark':8,'impact':'decrease','category':'channel'},
        {'name':'Abandon Rate','unit':'%','channels':['voice'],'processes':['all'],'benchmark':4,'impact':'decrease','category':'channel'},
        {'name':'Chat Concurrency','unit':'ratio','channels':['chat'],'processes':['all'],'benchmark':3.0,'impact':'increase','category':'channel'},
        {'name':'SLA %','unit':'%','channels':['voice','chat'],'processes':['all'],'benchmark':85,'impact':'increase','category':'channel'},
        {'name':'ACW','unit':'minutes','channels':['voice','chat'],'processes':['all'],'benchmark':1.5,'impact':'decrease','category':'channel'},
        {'name':'Voice Occupancy','unit':'%','channels':['voice'],'processes':['all'],'benchmark':82,'impact':'increase','category':'workforce'},
        {'name':'Chat Occupancy','unit':'%','channels':['chat'],'processes':['all'],'benchmark':80,'impact':'increase','category':'workforce'},
        {'name':'Shrinkage','unit':'%','channels':['all'],'processes':['all'],'benchmark':28,'impact':'decrease','category':'workforce'},
        {'name':'Schedule Efficiency','unit':'%','channels':['all'],'processes':['all'],'benchmark':92,'impact':'increase','category':'workforce'},
        {'name':'Attrition Rate','unit':'%','channels':['all'],'processes':['all'],'benchmark':3,'impact':'decrease','category':'workforce'},
        {'name':'Deflection Rate','unit':'%','channels':['portal'],'processes':['all'],'benchmark':30,'impact':'increase','category':'self_service'},
        {'name':'IVR Containment','unit':'%','channels':['voice'],'processes':['all'],'benchmark':25,'impact':'increase','category':'self_service'},
        {'name':'Repeat Contact Rate','unit':'%','channels':['voice','chat'],'processes':['orders'],'benchmark':12,'impact':'decrease','category':'industrial'},
    ]

# ═══════════════════════════════════════════════════════
# INITIATIVE LIBRARIES (enhanced with ramp + risk)
# ═══════════════════════════════════════════════════════
AUTOMATION_AI_LIBRARY = [
    {'name':'AI Agent Assist','levers':[{'lever':'aht_reduction','process_impacts':{'_all':0.12}}],'eligible_channels':['voice','chat'],'complexity':'medium','description':'Real-time AI suggestions during live conversations',
     'risk_category':'technology','risk_likelihood':0.3,'risk_impact':0.4},
    {'name':'Conversational Virtual Agent','levers':[{'lever':'deflection','process_impacts':{'_all':0.25}}],'eligible_channels':['voice','chat'],'complexity':'high','description':'AI chatbot/voicebot handling Tier 1 queries',
     'risk_category':'technology','risk_likelihood':0.4,'risk_impact':0.6},
    {'name':'GenAI Email Drafting','levers':[{'lever':'aht_reduction','process_impacts':{'_all':0.20}}],'eligible_channels':['email'],'complexity':'medium','description':'Auto-draft email responses for agent review',
     'risk_category':'technology','risk_likelihood':0.3,'risk_impact':0.3},
    {'name':'Auto-Summarization & Disposition','levers':[{'lever':'acw_reduction','process_impacts':{'_all':0.30}}],'eligible_channels':['voice','chat'],'complexity':'low','description':'AI-generated call summaries and auto-categorization',
     'risk_category':'technology','risk_likelihood':0.2,'risk_impact':0.2},
    {'name':'Semantic Knowledge Search','levers':[{'lever':'aht_reduction','process_impacts':{'_all':0.08}}],'eligible_channels':['voice','chat','email'],'complexity':'low','description':'AI-powered knowledge base search',
     'risk_category':'technology','risk_likelihood':0.2,'risk_impact':0.2},
    {'name':'Intelligent Call Routing','levers':[{'lever':'aht_reduction','process_impacts':{'_all':0.10}}],'eligible_channels':['voice','chat'],'complexity':'medium','description':'Skills-based routing using AI prediction',
     'risk_category':'technology','risk_likelihood':0.3,'risk_impact':0.3},
    {'name':'Predictive Customer Intent','levers':[{'lever':'aht_reduction','process_impacts':{'_all':0.07}}],'eligible_channels':['voice','chat'],'complexity':'high','description':'Predict reason for contact before connection',
     'risk_category':'technology','risk_likelihood':0.5,'risk_impact':0.3},
    {'name':'Self-Service Portal Enhancement','levers':[{'lever':'deflection','process_impacts':{'_all':0.15}}],'eligible_channels':['portal'],'complexity':'medium','description':'Enhanced self-service with AI guidance',
     'risk_category':'change','risk_likelihood':0.3,'risk_impact':0.4},
    {'name':'RPA for Back-Office Processing','levers':[{'lever':'aht_reduction','process_impacts':{'_all':0.15}}],'eligible_channels':['email','portal'],'complexity':'medium','description':'Automate repetitive back-office tasks',
     'risk_category':'technology','risk_likelihood':0.3,'risk_impact':0.3},
    {'name':'Proactive Notifications','levers':[{'lever':'deflection','process_impacts':{'_all':0.10}}],'eligible_channels':['voice','chat'],'complexity':'low','description':'Proactive alerts reduce inbound queries',
     'risk_category':'change','risk_likelihood':0.2,'risk_impact':0.3},
    {'name':'Speech Analytics','levers':[{'lever':'aht_reduction','process_impacts':{'_all':0.05}}],'eligible_channels':['voice'],'complexity':'medium','description':'Analyze calls for improvement patterns',
     'risk_category':'technology','risk_likelihood':0.2,'risk_impact':0.2},
    {'name':'Sentiment Analysis & Routing','levers':[{'lever':'aht_reduction','process_impacts':{'_all':0.06}}],'eligible_channels':['voice','chat'],'complexity':'medium','description':'Route by detected customer sentiment',
     'risk_category':'technology','risk_likelihood':0.3,'risk_impact':0.2},
    {'name':'Automated Quality Scoring','levers':[{'lever':'aht_reduction','process_impacts':{'_all':0.04}}],'eligible_channels':['voice','chat'],'complexity':'low','description':'AI-driven QA replacing manual review',
     'risk_category':'change','risk_likelihood':0.2,'risk_impact':0.2},
    {'name':'Visual IVR','levers':[{'lever':'deflection','process_impacts':{'_all':0.12}}],'eligible_channels':['voice'],'complexity':'medium','description':'Visual menu on mobile during IVR',
     'risk_category':'technology','risk_likelihood':0.3,'risk_impact':0.3},
    {'name':'AI Translation Assist','levers':[{'lever':'aht_reduction','process_impacts':{'_all':0.10}}],'eligible_channels':['chat','email'],'complexity':'low','description':'Real-time translation for multilingual support',
     'risk_category':'technology','risk_likelihood':0.2,'risk_impact':0.2},
    {'name':'Document OCR & Auto-Processing','levers':[{'lever':'aht_reduction','process_impacts':{'_all':0.18}}],'eligible_channels':['email','portal'],'complexity':'medium','description':'Auto-extract and process document data',
     'risk_category':'technology','risk_likelihood':0.3,'risk_impact':0.3},
    {'name':'Warranty Validation AI','levers':[{'lever':'aht_reduction','process_impacts':{'_all':0.15}}],'eligible_channels':['voice','chat','email'],'complexity':'medium','description':'AI validates warranty claims automatically',
     'risk_category':'technology','risk_likelihood':0.3,'risk_impact':0.3},
    {'name':'Parts Catalog AI Search','levers':[{'lever':'aht_reduction','process_impacts':{'_all':0.12}}],'eligible_channels':['voice','chat'],'complexity':'medium','description':'AI-powered parts identification and lookup',
     'risk_category':'technology','risk_likelihood':0.3,'risk_impact':0.2},
    {'name':'Repeat Contact Prevention','levers':[{'lever':'repeat_reduction','process_impacts':{'_all':0.20}}],'eligible_channels':['voice','chat'],'complexity':'medium','description':'Identify and prevent unnecessary callbacks',
     'risk_category':'process','risk_likelihood':0.3,'risk_impact':0.3},
    {'name':'Voice-to-Chat Shift','levers':[{'lever':'channel_shift','process_impacts':{'_all':0.15}}],'eligible_channels':['voice'],'complexity':'low','description':'Migrate voice contacts to chat channel',
     'source_channel':'voice','target_channel':'chat','risk_category':'change','risk_likelihood':0.4,'risk_impact':0.5},
]

OPMODEL_LIBRARY = [
    {'name':'Shrinkage Optimization','levers':[{'lever':'shrinkage_reduction','process_impacts':{'_all':0.15}}],'eligible_channels':['all'],'complexity':'low','description':'Reduce planned/unplanned absence through better WFM',
     'risk_category':'process','risk_likelihood':0.2,'risk_impact':0.2},
    {'name':'Schedule Optimization (WFM)','levers':[{'lever':'schedule_efficiency','process_impacts':{'_all':0.08}}],'eligible_channels':['all'],'complexity':'medium','description':'Better shift planning and rostering',
     'risk_category':'process','risk_likelihood':0.2,'risk_impact':0.3},
    {'name':'Occupancy Rebalancing','levers':[{'lever':'occupancy_uplift','process_impacts':{'_all':0.05}}],'eligible_channels':['voice','chat'],'complexity':'low','description':'Optimize agent utilization across queues',
     'risk_category':'process','risk_likelihood':0.2,'risk_impact':0.2},
    {'name':'Tiered Service Model','levers':[{'lever':'aht_reduction','process_impacts':{'_all':0.10}}],'eligible_channels':['voice','chat'],'complexity':'medium','description':'L1/L2/L3 restructure',
     'risk_category':'change','risk_likelihood':0.4,'risk_impact':0.5},
    {'name':'Cross-Skilling Program','levers':[{'lever':'occupancy_uplift','process_impacts':{'_all':0.06}}],'eligible_channels':['voice','chat'],'complexity':'medium','description':'Train agents across multiple queues',
     'risk_category':'change','risk_likelihood':0.3,'risk_impact':0.3},
    {'name':'Quality Framework Redesign','levers':[{'lever':'repeat_reduction','process_impacts':{'_all':0.10}}],'eligible_channels':['all'],'complexity':'medium','description':'Improve FCR through better quality standards',
     'risk_category':'process','risk_likelihood':0.3,'risk_impact':0.3},
    {'name':'Knowledge Management Overhaul','levers':[{'lever':'aht_reduction','process_impacts':{'_all':0.07}}],'eligible_channels':['all'],'complexity':'low','description':'Restructure knowledge base',
     'risk_category':'process','risk_likelihood':0.2,'risk_impact':0.2},
    {'name':'Process Standardization','levers':[{'lever':'aht_reduction','process_impacts':{'_all':0.08}}],'eligible_channels':['all'],'complexity':'medium','description':'Standardize handling procedures',
     'risk_category':'process','risk_likelihood':0.3,'risk_impact':0.3},
    {'name':'Concurrency Uplift Program','levers':[{'lever':'concurrency_uplift','process_impacts':{'_all':0.15}}],'eligible_channels':['chat'],'complexity':'low','description':'Handle more concurrent chats',
     'risk_category':'change','risk_likelihood':0.3,'risk_impact':0.3},
    {'name':'Workforce Consolidation','levers':[{'lever':'shrinkage_reduction','process_impacts':{'_all':0.10}}],'eligible_channels':['all'],'complexity':'high','description':'Consolidate smaller teams',
     'risk_category':'change','risk_likelihood':0.5,'risk_impact':0.6},
]

TECHNOLOGY_LIBRARY = [
    {'name':'CCaaS','category':'platform','cost_type':'both','one_time':100000,'recurring_monthly':8000,'description':'Cloud contact centre platform'},
    {'name':'Conversational AI Platform','category':'ai','cost_type':'both','one_time':80000,'recurring_monthly':6000,'description':'AI chatbot/voicebot engine'},
    {'name':'AI Agent Assist Tool','category':'ai','cost_type':'both','one_time':50000,'recurring_monthly':4000,'description':'Real-time AI guidance'},
    {'name':'GenAI Email Platform','category':'ai','cost_type':'both','one_time':40000,'recurring_monthly':3500,'description':'AI-powered email drafting'},
    {'name':'RPA Licenses','category':'automation','cost_type':'both','one_time':30000,'recurring_monthly':2000,'description':'Robotic process automation'},
    {'name':'WFM Platform','category':'wfm','cost_type':'both','one_time':60000,'recurring_monthly':5000,'description':'Workforce management system'},
    {'name':'Quality Management','category':'quality','cost_type':'both','one_time':35000,'recurring_monthly':3000,'description':'AI quality monitoring'},
    {'name':'Knowledge Management','category':'knowledge','cost_type':'both','one_time':25000,'recurring_monthly':2500,'description':'AI-powered knowledge base'},
    {'name':'Analytics & BI','category':'analytics','cost_type':'both','one_time':45000,'recurring_monthly':4000,'description':'Reporting and dashboards'},
    {'name':'Integration Middleware','category':'integration','cost_type':'both','one_time':40000,'recurring_monthly':3000,'description':'API gateway and connectors'},
    {'name':'Speech Analytics','category':'analytics','cost_type':'both','one_time':35000,'recurring_monthly':3500,'description':'Voice analysis platform'},
    {'name':'CRM Integration','category':'integration','cost_type':'both','one_time':50000,'recurring_monthly':2500,'description':'CRM connector and sync'},
]

# ═══════════════════════════════════════════════════════
# PROCESS COMPLEXITY (same as v6)
# ═══════════════════════════════════════════════════════
PROCESS_COMPLEXITY_DEFAULTS = {
    'orders': 0.35, 'warranty': 0.65, 'billing': 0.40,
    'parts': 0.50, 'technical_support': 0.70, 'dispatch': 0.55,
    'returns': 0.45, 'general': 0.35,
}

def _get_process_complexity(queue):
    override = queue.get('process_complexity')
    if override is not None: return float(override)
    ptag = queue.get('process_tag', 'general')
    return PROCESS_COMPLEXITY_DEFAULTS.get(ptag, 0.35)

def _complexity_band(pc):
    if pc <= 0.3: return 'simple'
    if pc <= 0.6: return 'medium'
    return 'complex'

STEP_LABELS = {'automation_ai':'Automation & AI','opmodel':'Operating Model','location':'Location Strategy'}

# ═══════════════════════════════════════════════════════
# INITIATIVE APPLICABILITY
# ═══════════════════════════════════════════════════════
def _init_applies_to_queue(init, queue):
    ch = queue['channel']
    ptag = queue.get('process_tag', 'general')
    elig_ch = init.get('eligible_channels', [])
    if elig_ch and 'all' not in elig_ch and ch not in elig_ch:
        return False
    levers = init.get('levers', [])
    if not levers: return True
    has_specific = False
    has_match = False
    for le in levers:
        pi = le.get('process_impacts', {})
        if '_all' in pi: has_match = True
        elif pi:
            has_specific = True
            if ptag in pi: has_match = True
    if not has_specific: return True
    return has_match

def _get_lever_impact_for_queue(lever_entry, queue):
    pi = lever_entry.get('process_impacts', {})
    ptag = queue.get('process_tag', 'general')
    if '_all' in pi: return pi['_all']
    return pi.get(ptag, 0)

# ═══════════════════════════════════════════════════════
# YEARLY RAMP WITH MONTHLY S-CURVE INTERPOLATION
# ═══════════════════════════════════════════════════════
def _ramp_at_month(init, month):
    """Yearly ramp targets with S-curve interpolation within each year.
    User sets rampYear1, rampYear2, rampYear3 (0-1).
    Within each year, we S-curve from previous year's target to current."""
    r1 = init.get('ramp_year1', 0.40)
    r2 = init.get('ramp_year2', 0.80)
    r3 = init.get('ramp_year3', 1.00)
    horizon = init.get('ramp_horizon', [r1, r2, r3])
    if isinstance(horizon, list) and len(horizon) >= 3:
        r1, r2, r3 = horizon[0], horizon[1], horizon[2]

    start = init.get('start_month', 1)
    if month < start: return 0

    year = min(((month - 1) // 12) + 1, 7)
    targets = [0, r1, r2, r3] + [1.0] * 4  # pad for years 4+

    prev_target = targets[year - 1] if year > 1 else 0
    curr_target = targets[min(year, len(targets) - 1)]

    # S-curve within the year
    month_in_year = ((month - 1) % 12) + 1
    t = month_in_year / 12.0
    s_curve = 1.0 / (1.0 + math.exp(-10 * (t - 0.5)))

    adoption_target = init.get('adoption_pct', 0.80)
    ramp = prev_target + (curr_target - prev_target) * s_curve
    return ramp * adoption_target


# ═══════════════════════════════════════════════════════
# BASELINE CALCULATION (same as v6)
# ═══════════════════════════════════════════════════════
def calc_queue_baseline(queue):
    ch = queue['channel']; vol = queue['monthly_volume']
    ht = queue['handle_time_minutes']; acw = queue.get('after_call_work_minutes', 0)
    conc = queue.get('chat_concurrency', 1.0) if ch == 'chat' else 1.0
    paid_hrs = queue.get('paid_hours_per_fte', 1920)
    shrink = queue.get('shrinkage_pct', 0.30); occ = queue.get('occupancy_target', 0.82)
    sched = queue.get('schedule_efficiency', 0.90)
    async_prod = queue.get('async_productivity_factor', 1.0) if ch in ['email','portal'] else 1.0
    total_ht_hrs = (ht + acw) / 60.0
    workload = (vol * total_ht_hrs) / max(conc, 1.0) if ch == 'chat' else vol * total_ht_hrs
    monthly_paid = paid_hrs / 12.0
    net_hrs = monthly_paid * (1 - shrink) * sched
    if ch in ['voice', 'chat']: fte = workload / max(net_hrs * occ, 1)
    else: fte = workload / max(net_hrs * async_prod, 1)
    return {'volume': vol, 'aht': ht, 'acw': acw, 'concurrency': conc, 'shrinkage': shrink,
            'occupancy': occ, 'schedule_efficiency': sched, 'fcr': queue.get('fcr_pct', 0.72),
            'transfer_rate': queue.get('transfer_pct', 0.08), 'fte': round(fte, 2),
            'workload_hours': round(workload, 2), 'net_productive_hours': round(net_hrs, 2)}


# ═══════════════════════════════════════════════════════
# HYBRID ENGINE: Pool + Stepped Realization
# ═══════════════════════════════════════════════════════
def apply_initiatives_pooled(queue, initiatives, month, pools, enterprise_pools):
    """Apply initiatives to a queue using pool-based netting with stepped realization.

    For each initiative × lever:
    1. Check pool ceiling — if exhausted, skip
    2. Apply stepped realization factor (90%/75%/60%/45%...)
    3. Consume from pool
    4. Apply the realized impact to queue state
    """
    ch = queue['channel']
    state = calc_queue_baseline(queue)
    vol = float(state['volume']); ht = float(state['aht']); acw = float(state['acw'])
    conc = float(state['concurrency']); shrink = float(state['shrinkage'])
    occ = float(state['occupancy']); sched = float(state['schedule_efficiency'])
    fcr = float(state['fcr']); transfer = float(state['transfer_rate'])
    repeat_pct = float(queue.get('repeat_contact_pct', 0.15))
    pc = _get_process_complexity(queue)

    deflected = 0; repeat_reduced = 0; shifted = 0
    init_contributions = []

    for init in initiatives:
        if not _init_applies_to_queue(init, queue): continue
        adoption = _ramp_at_month(init, month)
        if adoption <= 0: continue

        levers = init.get('levers', [])
        if not levers and init.get('lever'):
            old_procs = init.get('eligible_processes', ['all'])
            imp = init.get('impact_pct', 0)
            pi = {'_all': imp} if ('all' in old_procs or not old_procs) else {p: imp for p in old_procs}
            levers = [{'lever': init['lever'], 'process_impacts': pi}]

        for le in levers:
            lever = le.get('lever', '')
            impact_pct = _get_lever_impact_for_queue(le, queue)
            if impact_pct <= 0: continue

            # Process complexity modifier (lighter than v6 — just a discount, no curve)
            complexity_discount = 1.0 - pc * 0.3
            raw_impact = impact_pct * adoption * complexity_discount

            # Pool check: is there remaining capacity?
            pool = enterprise_pools.get(lever)
            if pool is None:
                continue  # FAIL CLOSED: unknown lever = 0 impact

            if pool['remaining_fte'] <= 0 and lever not in ['occupancy_uplift', 'schedule_efficiency']:
                continue  # pool exhausted

            # Stepped realization
            n = pool['initiatives_consuming']
            realization = _realization_factor(n)
            realized_impact = raw_impact * realization

            # Apply lever-specific logic
            fte_consumed = 0
            if lever == 'deflection':
                d = vol * realized_impact
                d = min(d, pool['remaining_contacts'])  # cap at pool
                deflected += d; vol -= d
                fte_consumed = d / max(state['volume'], 1) * state['fte']

            elif lever == 'repeat_reduction':
                rr = vol * repeat_pct * realized_impact
                rr = min(rr, pool['remaining_contacts'])
                repeat_reduced += rr; vol -= rr
                fcr = min(fcr + realized_impact * 0.5, 0.98)
                fte_consumed = rr / max(state['volume'], 1) * state['fte']

            elif lever == 'channel_shift':
                if init.get('source_channel') == ch:
                    s = vol * realized_impact
                    s = min(s, pool['remaining_contacts'])
                    shifted += s; vol -= s
                    fte_consumed = s / max(state['volume'], 1) * state['fte']

            elif lever == 'aht_reduction':
                old_ht = ht
                ht = max(ht * (1 - realized_impact), ht * 0.40)  # floor at 40% of current
                fte_consumed = state['fte'] * (old_ht - ht) / max(old_ht, 0.1)

            elif lever == 'acw_reduction':
                old_acw = acw
                acw = max(acw * (1 - realized_impact), 0)
                # FTE impact proportional to ACW share of total handle time
                acw_share = old_acw / max(state['aht'] + old_acw, 1)
                fte_consumed = state['fte'] * (old_acw - acw) / max(old_acw, 0.1) * acw_share

            elif lever == 'concurrency_uplift':
                if ch == 'chat':
                    old_conc = conc
                    conc = min(conc * (1 + realized_impact), 6.0)
                    fte_consumed = state['fte'] * (1 - old_conc / max(conc, 1))

            elif lever == 'shrinkage_reduction':
                old_shrink = shrink
                shrink = max(shrink * (1 - realized_impact), 0.12)
                fte_consumed = state['fte'] * (old_shrink - shrink) / max(1 - old_shrink, 0.1)

            elif lever == 'occupancy_uplift':
                ceil = 0.92 if ch == 'voice' else (0.90 if ch == 'chat' else 0.95)
                old_occ = occ
                occ = min(occ * (1 + realized_impact), ceil)
                fte_consumed = state['fte'] * (occ - old_occ) / max(occ, 0.1)

            elif lever == 'schedule_efficiency':
                old_sched = sched
                sched = min(sched * (1 + realized_impact), 0.98)
                fte_consumed = state['fte'] * (sched - old_sched) / max(sched, 0.1)

            # Consume from pool
            fte_consumed = max(fte_consumed, 0)
            if pool['ceiling_fte'] > 0:
                fte_consumed = min(fte_consumed, pool['remaining_fte'])
            pool['consumed_fte'] = round(pool.get('consumed_fte', 0) + fte_consumed, 2)
            pool['remaining_fte'] = round(pool.get('ceiling_fte', 0) - pool.get('consumed_fte', 0), 2)
            if lever in ['deflection', 'repeat_reduction', 'channel_shift']:
                consumed_contacts = d if lever == 'deflection' else (rr if lever == 'repeat_reduction' else (s if lever == 'channel_shift' else 0))
                pool['consumed_contacts'] = round(pool.get('consumed_contacts', 0) + consumed_contacts, 0)
                pool['remaining_contacts'] = round(pool.get('ceiling_contacts', 0) - pool.get('consumed_contacts', 0), 0)
            pool['initiatives_consuming'] = pool.get('initiatives_consuming', 0) + 1

    # Recalculate FTE from new state
    total_ht_hrs = (ht + acw) / 60.0
    workload = (vol * total_ht_hrs) / max(conc, 1.0) if ch == 'chat' else vol * total_ht_hrs
    paid_hrs = queue.get('paid_hours_per_fte', 1920); monthly_paid = paid_hrs / 12.0
    net_hrs = monthly_paid * (1 - shrink) * sched
    async_prod = queue.get('async_productivity_factor', 1.0) if ch in ['email', 'portal'] else 1.0
    if ch in ['voice', 'chat']:
        fte = workload / max(net_hrs * occ, 1)
    else:
        fte = workload / max(net_hrs * async_prod, 1)

    return {'volume': round(vol, 0), 'aht': round(ht, 2), 'acw': round(acw, 2),
            'concurrency': round(conc, 2), 'shrinkage': round(shrink, 4),
            'occupancy': round(occ, 4), 'schedule_efficiency': round(sched, 4),
            'fcr': round(fcr, 4),
            'transfer_rate': round(transfer * (1 - (state['aht'] - ht) / max(state['aht'], 1) * 0.3), 4),
            'fte': round(fte, 2), 'deflected': round(deflected, 0),
            'repeat_reduced': round(repeat_reduced, 0), 'shifted': round(shifted, 0),
            'workload_hours': round(workload, 2)}


# ═══════════════════════════════════════════════════════
# BLENDED COST & LOCATION (same as v6)
# ═══════════════════════════════════════════════════════
def _calc_blended_cost(enterprise):
    lm = enterprise.get('location_mix', {})
    sr = enterprise.get('salary_rates', {})
    total_fte = sum(v.get('fte', 0) for v in lm.values())
    if total_fte <= 0:
        rates = [v for v in sr.values() if v > 0]
        return sum(rates) / max(len(rates), 1) if rates else 55000
    blended = 0
    for loc, info in lm.items():
        fte = info.get('fte', 0)
        if fte > 0:
            pct = fte / total_fte
            blended += pct * sr.get(loc, 55000)
    return round(blended, 2)

def _distribute_fte_by_location(fte, enterprise):
    lm = enterprise.get('location_mix', {})
    total_mix = sum(v.get('fte', 0) for v in lm.values())
    dist = {}
    if total_mix > 0:
        for loc, info in lm.items():
            mix_fte = info.get('fte', 0)
            if mix_fte > 0:
                dist[loc] = round(fte * (mix_fte / total_mix), 2)
    else:
        dist['onshore'] = round(fte, 2)
    return dist

def apply_location_savings(queue_state, location_moves, queue, salary_rates, bu_cost, enterprise=None):
    fte = queue_state['fte']
    if enterprise:
        fte_before = _distribute_fte_by_location(fte, enterprise)
        blended_before = _calc_blended_cost(enterprise)
    else:
        fte_before = {'onshore': fte}
        blended_before = bu_cost
    cost_before = fte * blended_before / 12
    fte_after = dict(fte_before)
    for move in location_moves:
        mp = move.get('processes', ['all']); mc = move.get('channels', ['all'])
        if 'all' not in mp and queue.get('process_tag', '') not in mp: continue
        if 'all' not in mc and queue['channel'] not in mc: continue
        from_loc = move.get('from_location', 'onshore')
        to_loc = move.get('to_location', 'nearshore')
        move_pct = move.get('move_pct', 0)
        available = fte_after.get(from_loc, 0)
        moved = round(available * move_pct, 2)
        fte_after[from_loc] = round(max(available - moved, 0), 2)
        fte_after[to_loc] = round(fte_after.get(to_loc, 0) + moved, 2)
    total_after = sum(fte_after.values())
    if total_after > 0:
        blended_after = sum((fte_after.get(loc, 0) / total_after) * salary_rates.get(loc, 55000)
                            for loc in fte_after if fte_after.get(loc, 0) > 0)
    else:
        blended_after = blended_before
    cost_after = fte * blended_after / 12
    fte_before = {k: v for k, v in fte_before.items() if v > 0}
    fte_after = {k: v for k, v in fte_after.items() if v > 0}
    return {'fte': fte, 'cost_before': round(cost_before, 0), 'cost_after': round(cost_after, 0),
            'location_saving': round(cost_before - cost_after, 0),
            'effective_cost_per_fte': round(blended_after, 0),
            'blended_before': round(blended_before, 0),
            'fte_by_location_before': fte_before, 'fte_by_location_after': fte_after}


# ═══════════════════════════════════════════════════════
# FINANCIALS (IRR, CX Revenue, Cost of Inaction, Sensitivity, Scenarios)
# ═══════════════════════════════════════════════════════
def _calc_irr(cash_flows, max_iter=200, tol=1e-7):
    if not cash_flows or all(c == 0 for c in cash_flows): return 0
    r = 0.10
    for _ in range(max_iter):
        npv = sum(cf / (1 + r) ** t for t, cf in enumerate(cash_flows))
        dnpv = sum(-t * cf / (1 + r) ** (t + 1) for t, cf in enumerate(cash_flows))
        if abs(dnpv) < 1e-12: break
        r_new = r - npv / dnpv
        if abs(r_new - r) < tol: return round(r_new * 100, 1)
        r = r_new
    return round(r * 100, 1)

def _calc_cost_of_inaction(enterprise, horizon_years):
    growth = enterprise.get('global_volume_growth_pct', 0.02)
    inflation = enterprise.get('global_wage_inflation_pct', 0.03)
    attrition_m = enterprise.get('attrition_rate_monthly', 0.03)
    enterprise_fte = enterprise.get('enterprise_fte', 0)
    blended_cost = _calc_blended_cost(enterprise)
    hiring_cost_per_fte = 5000
    results = []
    for year in range(1, horizon_years + 1):
        vol_factor = (1 + growth) ** year
        infl_factor = (1 + inflation) ** year
        fte_needed = enterprise_fte * vol_factor
        annual_cost = fte_needed * blended_cost * infl_factor
        annual_attrition = fte_needed * (1 - (1 - attrition_m) ** 12)
        attrition_cost = annual_attrition * hiring_cost_per_fte
        total = annual_cost + attrition_cost
        results.append({'year': year, 'fte_needed': round(fte_needed, 1),
                        'labor_cost': round(annual_cost, 0), 'attrition_cost': round(attrition_cost, 0),
                        'total_cost': round(total, 0)})
    return results

def _calc_cx_revenue_impact(enterprise, horizon_years):
    cx = enterprise.get('cx_revenue', {})
    customer_base = cx.get('customer_base', 0)
    rev_per_customer = cx.get('revenue_per_customer', 0)
    baseline_churn = cx.get('annual_churn_rate_pct', 0.12)
    csat_baseline = cx.get('csat_baseline', 3.5)
    csat_target = cx.get('csat_target', 4.2)
    fcr_baseline = cx.get('fcr_baseline_pct', 0.70)
    fcr_target = cx.get('fcr_target_pct', 0.82)
    if customer_base <= 0 or rev_per_customer <= 0:
        return {'enabled': False, 'yearly': []}
    csat_improvement = max(0, csat_target - csat_baseline)
    fcr_improvement = max(0, fcr_target - fcr_baseline)
    churn_reduction_from_csat = csat_improvement * 0.015
    churn_reduction_from_fcr = fcr_improvement * 0.005
    total_churn_reduction = min(churn_reduction_from_csat + churn_reduction_from_fcr, baseline_churn * 0.5)
    new_churn = baseline_churn - total_churn_reduction
    yearly = []
    for year in range(1, horizon_years + 1):
        ramp = min(1.0, year * 0.4)
        effective_reduction = total_churn_reduction * ramp
        retained_customers = round(customer_base * effective_reduction)
        revenue_retained = round(retained_customers * rev_per_customer, 0)
        yearly.append({'year': year, 'retained_customers': retained_customers,
                       'revenue_retained': revenue_retained,
                       'churn_rate': round((baseline_churn - effective_reduction) * 100, 1)})
    return {'enabled': True, 'baseline_churn_pct': round(baseline_churn * 100, 1),
            'projected_churn_pct': round(new_churn * 100, 1),
            'churn_reduction_pct': round(total_churn_reduction * 100, 2),
            'yearly': yearly, 'total_revenue_retained': sum(y['revenue_retained'] for y in yearly)}

def _run_sensitivity(enterprise, base_result):
    if not base_result or 'summary' not in base_result: return []
    base_npv = base_result['summary'].get('npv', 0)
    variables = [
        ('volume_growth', 'global_volume_growth_pct', 'Volume Growth'),
        ('wage_inflation', 'global_wage_inflation_pct', 'Wage Inflation'),
        ('discount_rate', 'discount_rate', 'Discount Rate'),
        ('attrition', 'attrition_rate_monthly', 'Attrition Rate'),
        ('redeployment', 'redeployment_pct', 'Redeployment %'),
    ]
    results = []
    for key, field, label in variables:
        original = enterprise.get(field, 0)
        if original == 0: continue
        for direction, mult in [('low', 0.8), ('high', 1.2)]:
            mod = copy.deepcopy(enterprise)
            mod[field] = original * mult
            try:
                r = run_waterfall(mod)
                npv = r.get('summary', {}).get('npv', 0)
            except:
                npv = base_npv
            results.append({'variable': label, 'direction': direction,
                           'value': round(original * mult, 4), 'npv': round(npv, 0)})
    tornado = []
    seen = set()
    for item in results:
        if item['variable'] not in seen:
            low = next((r for r in results if r['variable'] == item['variable'] and r['direction'] == 'low'), None)
            high = next((r for r in results if r['variable'] == item['variable'] and r['direction'] == 'high'), None)
            if low and high:
                tornado.append({'variable': item['variable'],
                               'low_npv': low['npv'], 'high_npv': high['npv'],
                               'base_npv': round(base_npv, 0),
                               'swing': abs(high['npv'] - low['npv'])})
            seen.add(item['variable'])
    tornado.sort(key=lambda x: x['swing'], reverse=True)
    return tornado

def _run_scenario_comparison(enterprise, base_result):
    scenarios = enterprise.get('scenario_multipliers', {})
    if not scenarios: return {}
    results = {}
    for scenario_name, mults in scenarios.items():
        if scenario_name == 'base':
            results['base'] = {
                'npv': base_result['summary']['npv'],
                'roi_pct': base_result['summary']['roi_pct'],
                'payback_year': base_result['summary']['payback_year'],
                'fte_reduction': base_result['summary']['fte_reduction'],
                'total_net_benefit': base_result['summary']['total_net_benefit'],
                'yearly_net': [y['net_saving'] for y in base_result.get('yearly_data', [])],
            }
            continue
        mod = copy.deepcopy(enterprise)
        adoption_mult = mults.get('adoption', 1.0)
        for init_list in ['initiatives_auto', 'initiatives_opmodel']:
            for init in mod.get(init_list, []):
                for rk in ['ramp_year1', 'ramp_year2', 'ramp_year3']:
                    if rk in init:
                        init[rk] = min(1.0, init[rk] * adoption_mult)
        vol_mult = mults.get('volume_growth', 1.0)
        mod['global_volume_growth_pct'] = enterprise.get('global_volume_growth_pct', 0.02) * vol_mult
        try:
            r = run_waterfall(mod)
            s = r.get('summary', {})
            ben_mult = mults.get('benefit_realization', 1.0)
            results[scenario_name] = {
                'npv': round(s.get('npv', 0) * ben_mult, 0),
                'roi_pct': round(s.get('roi_pct', 0) * ben_mult, 1),
                'payback_year': s.get('payback_year'),
                'fte_reduction': s.get('fte_reduction', 0),
                'total_net_benefit': round(s.get('total_net_benefit', 0) * ben_mult, 0),
                'yearly_net': [round(y['net_saving'] * ben_mult, 0) for y in r.get('yearly_data', [])],
            }
        except:
            results[scenario_name] = {'npv': 0, 'roi_pct': 0, 'payback_year': None,
                                       'fte_reduction': 0, 'total_net_benefit': 0, 'yearly_net': []}
    return results


# ═══════════════════════════════════════════════════════
# ROLE-LEVEL FTE BREAKDOWN
# ═══════════════════════════════════════════════════════
def _calc_role_breakdown(enterprise, baseline_fte, final_fte):
    """Distribute FTE reduction across roles based on role distribution and reducibility."""
    roles = enterprise.get('roles', ROLE_DEFAULTS)
    total_reduction = baseline_fte - final_fte
    reducible_total_pct = sum(r['pct_of_fte'] for r in roles if r.get('reducible', True))
    blended_cost = _calc_blended_cost(enterprise)
    breakdown = []
    for role in roles:
        baseline = round(baseline_fte * role['pct_of_fte'], 1)
        if role.get('reducible', True) and reducible_total_pct > 0:
            share = role['pct_of_fte'] / reducible_total_pct
            reduction = round(total_reduction * share, 1)
        else:
            reduction = 0
        future = round(max(baseline - reduction, 0), 1)
        cost_per = blended_cost * role.get('cost_multiplier', 1.0)
        saving = round(reduction * cost_per, 0)
        breakdown.append({
            'role': role['role'],
            'baseline_fte': baseline,
            'future_fte': future,
            'reduction': reduction,
            'pct_reduction': round(reduction / max(baseline, 1) * 100, 1),
            'cost_per_fte': round(cost_per, 0),
            'annual_saving': saving,
        })
    return breakdown


# ═══════════════════════════════════════════════════════
# RISK REGISTER
# ═══════════════════════════════════════════════════════
def _generate_risk_register(enterprise, all_inits):
    """Auto-generate risk register from initiative profiles."""
    risks = []
    risk_id = 1
    category_risks = {}
    for init in all_inits:
        cat = init.get('risk_category', 'general')
        likelihood = init.get('risk_likelihood', 0.3)
        impact = init.get('risk_impact', 0.3)
        score = round(likelihood * impact * 100, 0)
        risk = {
            'id': f'R{risk_id:03d}',
            'initiative': init.get('name', ''),
            'category': cat,
            'description': _risk_description(init),
            'likelihood': likelihood,
            'impact': impact,
            'score': score,
            'rating': 'high' if score > 20 else ('medium' if score > 10 else 'low'),
            'mitigation': _risk_mitigation(cat, init.get('complexity', 'medium')),
        }
        risks.append(risk)
        risk_id += 1
        if cat not in category_risks:
            category_risks[cat] = []
        category_risks[cat].append(score)

    # Category summary
    category_summary = {}
    for cat, scores in category_risks.items():
        category_summary[cat] = {
            'count': len(scores),
            'avg_score': round(sum(scores) / max(len(scores), 1), 1),
            'max_score': max(scores) if scores else 0,
        }

    risks.sort(key=lambda r: -r['score'])
    return {
        'risks': risks[:20],
        'category_summary': category_summary,
        'total_risks': len(risks),
        'high_risks': sum(1 for r in risks if r['rating'] == 'high'),
        'medium_risks': sum(1 for r in risks if r['rating'] == 'medium'),
        'low_risks': sum(1 for r in risks if r['rating'] == 'low'),
    }

def _risk_description(init):
    complexity = init.get('complexity', 'medium')
    name = init.get('name', '')
    if complexity == 'high':
        return f"{name}: High implementation complexity may delay benefits realization and require significant change management"
    elif complexity == 'medium':
        return f"{name}: Moderate complexity requiring careful planning and stakeholder alignment"
    return f"{name}: Lower complexity but requires consistent execution and monitoring"

def _risk_mitigation(category, complexity):
    mitigations = {
        'technology': 'Phased rollout with pilot, vendor SLA guarantees, fallback plan',
        'change': 'Change champion network, communication plan, staged adoption targets',
        'process': 'Process mapping workshops, baseline measurement, continuous improvement cycle',
        'general': 'Regular governance reviews, clear ownership, escalation framework',
    }
    base = mitigations.get(category, mitigations['general'])
    if complexity == 'high':
        base += '; dedicated PMO oversight'
    return base


# ═══════════════════════════════════════════════════════
# ATTRITION TIMELINE
# ═══════════════════════════════════════════════════════
def _calc_attrition_timeline(enterprise, total_reduction):
    """Month-by-month natural attrition vs required reduction."""
    attrition_m = enterprise.get('attrition_rate_monthly', 0.03)
    enterprise_fte = enterprise.get('enterprise_fte', 0)
    if enterprise_fte <= 0 or total_reduction <= 0:
        return {'months': [], 'crossover_month': None}
    months = []
    crossover = None
    for m in range(1, 37):
        cumulative_attrited = enterprise_fte * (1 - (1 - attrition_m) ** m)
        months.append({
            'month': m,
            'cumulative_attrition': round(cumulative_attrited, 1),
            'required_reduction': round(total_reduction, 1),
            'gap': round(total_reduction - cumulative_attrited, 1),
        })
        if crossover is None and cumulative_attrited >= total_reduction:
            crossover = m
    return {'months': months, 'crossover_month': crossover}


# ═══════════════════════════════════════════════════════
# MAIN WATERFALL ENGINE
# ═══════════════════════════════════════════════════════
def run_waterfall(enterprise):
    horizon_years = enterprise.get('planning_horizon_years', 3)
    horizon_months = horizon_years * 12
    salary_rates = enterprise.get('salary_rates', {})
    growth_pct = enterprise.get('global_volume_growth_pct', 0.02)
    inflation_pct = enterprise.get('global_wage_inflation_pct', 0.03)
    attrition = enterprise.get('attrition_rate_monthly', 0.03)
    redeployment = enterprise.get('redeployment_pct', 0.10)
    discount_rate = enterprise.get('discount_rate', 0.10)
    enterprise_fte = enterprise.get('enterprise_fte', 0)
    waterfall_order = enterprise.get('waterfall_order', ['automation_ai', 'opmodel', 'location'])
    waterfall_active = enterprise.get('waterfall_active', {'automation_ai': True, 'opmodel': True, 'location': True})
    active_steps = [s for s in waterfall_order if waterfall_active.get(s, True)]

    auto_inits = enterprise.get('initiatives_auto', []) if 'automation_ai' in active_steps else []
    opmodel_inits = enterprise.get('initiatives_opmodel', []) if 'opmodel' in active_steps else []
    location_moves = enterprise.get('location_strategy', []) if 'location' in active_steps else []
    tech_items = enterprise.get('technology', [])

    blended_baseline_cost = _calc_blended_cost(enterprise)

    all_queues = []
    for bu in enterprise.get('business_units', []):
        for q in bu.get('queues', []):
            all_queues.append((bu, q))
    if not all_queues:
        return {'error': 'No queues defined.'}

    if enterprise_fte <= 0:
        enterprise_fte = sum(bu.get('current_fte', 0) for bu in enterprise.get('business_units', []))
    if enterprise_fte <= 0:
        enterprise_fte = sum(calc_queue_baseline(q)['fte'] for _, q in all_queues)

    # Compute baselines and scale to enterprise FTE
    raw_baselines = []
    total_calc_fte = 0
    for bu, q in all_queues:
        bl = calc_queue_baseline(q); raw_baselines.append(bl); total_calc_fte += bl['fte']
    fte_scale = enterprise_fte / max(total_calc_fte, 0.1) if total_calc_fte > 0 else 1.0
    fixed_baseline_ftes = [round(raw_baselines[idx]['fte'] * fte_scale, 2) for idx in range(len(all_queues))]

    # Compute pool ceilings (enterprise-wide)
    enterprise_pools = compute_pool_ceilings(all_queues, enterprise)
    # Scale pool ceilings to enterprise FTE
    for lever, pool in enterprise_pools.items():
        pool['ceiling_fte'] = round(pool['ceiling_fte'] * fte_scale, 2)
        pool['remaining_fte'] = pool['ceiling_fte']

    yearly_data = []
    queue_results = []
    pool_snapshots = []

    for year in range(1, horizon_years + 1):
        month = year * 12
        year_queues = []
        growth_factor = (1 + growth_pct) ** year
        inflation_factor = (1 + inflation_pct) ** year

        # Reset pools per year (initiatives accumulate across queues but reset each year for yearly reporting)
        year_pools = copy.deepcopy(enterprise_pools)

        for idx, (bu, q) in enumerate(all_queues):
            grown_q = copy.deepcopy(q)
            grown_q['monthly_volume'] = round(q['monthly_volume'] * growth_factor)
            baseline = calc_queue_baseline(q)
            scaled_baseline_fte = fixed_baseline_ftes[idx]
            baseline['fte'] = scaled_baseline_fte

            step_results = {}
            prev_state = baseline
            prev_fte = scaled_baseline_fte
            effective_cost = blended_baseline_cost

            for step in waterfall_order:
                if step not in active_steps: continue

                if step == 'automation_ai':
                    # Build queue with previous state for this step
                    vq = copy.deepcopy(grown_q)
                    vq['monthly_volume'] = prev_state['volume']
                    vq['handle_time_minutes'] = prev_state['aht']
                    vq['after_call_work_minutes'] = prev_state['acw']
                    if q['channel'] == 'chat': vq['chat_concurrency'] = prev_state['concurrency']
                    vq['shrinkage_pct'] = prev_state['shrinkage']
                    vq['occupancy_target'] = prev_state['occupancy']
                    vq['schedule_efficiency'] = prev_state['schedule_efficiency']
                    vq['fcr_pct'] = prev_state['fcr']

                    r = apply_initiatives_pooled(vq, auto_inits, month, year_pools, year_pools)
                    # Scale FTE
                    vq_baseline = calc_queue_baseline(vq)
                    ratio = r['fte'] / vq_baseline['fte'] if vq_baseline['fte'] > 0 else 1.0
                    r['fte'] = round(prev_fte * ratio, 2)
                    step_results['automation_ai'] = r
                    prev_state = r; prev_fte = r['fte']

                elif step == 'opmodel':
                    vq = copy.deepcopy(grown_q)
                    vq['monthly_volume'] = prev_state['volume']
                    vq['handle_time_minutes'] = prev_state['aht']
                    vq['after_call_work_minutes'] = prev_state['acw']
                    if q['channel'] == 'chat': vq['chat_concurrency'] = prev_state['concurrency']
                    vq['shrinkage_pct'] = prev_state['shrinkage']
                    vq['occupancy_target'] = prev_state['occupancy']
                    vq['schedule_efficiency'] = prev_state['schedule_efficiency']
                    vq['fcr_pct'] = prev_state['fcr']

                    r = apply_initiatives_pooled(vq, opmodel_inits, month, year_pools, year_pools)
                    vq_baseline = calc_queue_baseline(vq)
                    ratio = r['fte'] / vq_baseline['fte'] if vq_baseline['fte'] > 0 else 1.0
                    r['fte'] = round(prev_fte * ratio, 2)
                    step_results['opmodel'] = r
                    prev_state = r; prev_fte = r['fte']

                elif step == 'location':
                    loc = apply_location_savings(prev_state, location_moves, q, salary_rates,
                                                  blended_baseline_cost, enterprise)
                    step_results['location'] = loc
                    effective_cost = loc['effective_cost_per_fte']

            baseline_cost_monthly = baseline['fte'] * blended_baseline_cost * inflation_factor / 12
            final_cost_monthly = prev_fte * effective_cost * inflation_factor / 12

            year_queues.append({
                'bu': bu['bu_name'], 'queue': q['queue_name'], 'channel': q['channel'],
                'process': q.get('process_tag', ''), 'year': year, 'baseline': baseline,
                'step_results': step_results, 'baseline_cost_monthly': round(baseline_cost_monthly, 0),
                'final_cost_monthly': round(final_cost_monthly, 0),
                'waterfall_order': waterfall_order, 'active_steps': active_steps
            })

        # Aggregate year
        bl_fte = enterprise_fte
        step_fte_totals = {}
        for step in active_steps:
            if step == 'location':
                prev_step_fte = bl_fte
                for s in active_steps:
                    if s == 'location': break
                    if s in step_fte_totals: prev_step_fte = step_fte_totals[s]
                step_fte_totals[step] = prev_step_fte
            else:
                step_fte_totals[step] = sum(
                    yq['step_results'].get(step, yq['baseline']).get('fte', 0) for yq in year_queues)

        final_fte = bl_fte
        for s in reversed(active_steps):
            if s != 'location' and s in step_fte_totals:
                final_fte = step_fte_totals[s]; break

        bl_cost = sum(yq['baseline_cost_monthly'] for yq in year_queues) * 12
        final_cost = sum(yq['final_cost_monthly'] for yq in year_queues) * 12
        labor_saving = bl_cost - final_cost

        avg_cost_per_fte = bl_cost / max(bl_fte, 1)
        step_savings = {}; pfte = bl_fte
        for step in active_steps:
            if step == 'location':
                step_savings[step] = labor_saving - sum(step_savings.get(s, 0) for s in active_steps if s != 'location')
            else:
                sfte = step_fte_totals.get(step, bl_fte)
                step_savings[step] = (pfte - sfte) * avg_cost_per_fte
                pfte = sfte

        # Technology costs
        tech_one_time = 0; tech_recurring = 0
        for t in tech_items:
            sm = t.get('start_month', 1); em = t.get('end_month', 36)
            ys = (year - 1) * 12 + 1; ye = year * 12
            if sm <= ye and em >= ys:
                if year == 1 or (sm >= ys and sm <= ye):
                    tech_one_time += t.get('one_time', 0)
                tech_recurring += t.get('recurring_monthly', 0) * max(min(ye, em) - max(ys, sm) + 1, 0)
        total_tech = tech_one_time + tech_recurring

        attrition_factor = min(1.0, 1 - (1 - attrition) ** (year * 12))
        realized_saving = labor_saving * attrition_factor
        redeployed_cost = realized_saving * redeployment
        net_saving = realized_saving - redeployed_cost - total_tech

        # Pool snapshot for this year
        pool_snapshots.append({
            'year': year,
            'pools': {lever: {
                'ceiling_fte': pool['ceiling_fte'],
                'consumed_fte': pool['consumed_fte'],
                'remaining_fte': pool['remaining_fte'],
                'utilization_pct': round(pool['consumed_fte'] / max(pool['ceiling_fte'], 0.1) * 100, 1),
                'initiatives_consuming': pool['initiatives_consuming'],
            } for lever, pool in year_pools.items() if pool['ceiling_fte'] > 0}
        })

        yearly_data.append({
            'year': year, 'baseline_fte': round(bl_fte, 1), 'final_fte': round(final_fte, 1),
            'fte_reduction': round(bl_fte - final_fte, 1),
            'baseline_cost': round(bl_cost, 0), 'final_cost': round(final_cost, 0),
            'total_labor_saving': round(labor_saving, 0),
            'step_fte': {s: round(step_fte_totals.get(s, bl_fte), 1) for s in active_steps},
            'step_savings': {s: round(step_savings.get(s, 0), 0) for s in active_steps},
            'active_steps': active_steps,
            'tech_one_time': round(tech_one_time, 0), 'tech_recurring': round(tech_recurring, 0),
            'total_tech': round(total_tech, 0),
            'attrition_factor': round(attrition_factor, 4), 'realized_saving': round(realized_saving, 0),
            'redeployed_cost': round(redeployed_cost, 0), 'net_saving': round(net_saving, 0),
            'waterfall_order': waterfall_order, 'queue_details': year_queues
        })
        queue_results.extend(year_queues)

    # Summary
    total_labor_saving = sum(y['total_labor_saving'] for y in yearly_data)
    total_tech_cost = sum(y['total_tech'] for y in yearly_data)
    total_net = sum(y['net_saving'] for y in yearly_data)
    cumulative = []; running = 0
    for y in yearly_data: running += y['net_saving']; cumulative.append(round(running, 0))
    payback_year = None
    for i, c in enumerate(cumulative):
        if c > 0: payback_year = i + 1; break
    roi_pct = round((total_net / max(total_tech_cost, 1)) * 100, 1) if total_tech_cost > 0 else 0
    npv = sum(y['net_saving'] / ((1 + discount_rate) ** y['year']) for y in yearly_data)

    # IRR
    impl = enterprise.get('implementation_costs', {})
    impl_total = sum(impl.get(k, 0) for k in ['change_management', 'training', 'integration'])
    contingency = impl_total * impl.get('contingency_pct', 0.10)
    impl_grand_total = impl_total + contingency
    yr0_investment = -(total_tech_cost + impl_grand_total) if yearly_data else 0
    cash_flows = [yr0_investment] + [y['net_saving'] for y in yearly_data]
    irr = _calc_irr(cash_flows)

    # Additional analytics
    cost_of_inaction = _calc_cost_of_inaction(enterprise, horizon_years)
    cx_revenue_impact = _calc_cx_revenue_impact(enterprise, horizon_years)
    impl_breakdown = {
        'change_management': impl.get('change_management', 0),
        'training': impl.get('training', 0),
        'integration': impl.get('integration', 0),
        'contingency': round(contingency, 0),
        'technology_total': round(total_tech_cost, 0),
        'grand_total': round(total_tech_cost + impl_grand_total, 0),
    }

    # Initiative impacts
    initiative_impacts = calc_initiative_impacts(enterprise, all_queues, horizon_months, active_steps,
                                                  enterprise_fte, fte_scale, fixed_baseline_ftes)

    # KPI gaps
    kpi_gaps = calc_kpi_gaps(enterprise)
    recs = generate_recommendations(kpi_gaps, enterprise)

    # Validation
    validation = run_validation(enterprise, all_queues, yearly_data, active_steps)

    # Role breakdown
    final_fte_val = yearly_data[-1]['final_fte'] if yearly_data else 0
    role_breakdown = _calc_role_breakdown(enterprise, enterprise_fte, final_fte_val)

    # Risk register
    all_inits = enterprise.get('initiatives_auto', []) + enterprise.get('initiatives_opmodel', [])
    risk_register = _generate_risk_register(enterprise, all_inits)

    # Attrition timeline
    total_reduction = enterprise_fte - final_fte_val
    attrition_timeline = _calc_attrition_timeline(enterprise, total_reduction)

    # Location breakdown
    loc_before_agg = {}; loc_after_agg = {}
    if yearly_data:
        for yq in yearly_data[-1].get('queue_details', []):
            loc_res = yq.get('step_results', {}).get('location', {})
            fb = loc_res.get('fte_by_location_before', {})
            fa = loc_res.get('fte_by_location_after', {})
            for loc, fte in fb.items(): loc_before_agg[loc] = round(loc_before_agg.get(loc, 0) + fte, 1)
            for loc, fte in fa.items(): loc_after_agg[loc] = round(loc_after_agg.get(loc, 0) + fte, 1)
    if not loc_before_agg:
        loc_before_agg = _distribute_fte_by_location(enterprise_fte, enterprise)
        loc_after_agg = dict(loc_before_agg)
    location_breakdown = {
        'before': loc_before_agg, 'after': loc_after_agg,
        'blended_cost_before': round(blended_baseline_cost, 0),
        'blended_cost_after': round(
            yearly_data[-1]['queue_details'][0]['step_results'].get('location', {}).get(
                'effective_cost_per_fte', blended_baseline_cost), 0
        ) if yearly_data and yearly_data[-1].get('queue_details') else round(blended_baseline_cost, 0)
    }

    return {
        'summary': {
            'baseline_fte': round(enterprise_fte, 1),
            'final_fte': final_fte_val,
            'fte_reduction': round(enterprise_fte - final_fte_val, 1),
            'total_labor_saving': round(total_labor_saving, 0),
            'total_tech_cost': round(total_tech_cost, 0),
            'total_net_benefit': round(total_net, 0),
            'roi_pct': roi_pct, 'payback_year': payback_year,
            'npv': round(npv, 0), 'irr': irr,
            'horizon_years': horizon_years, 'cumulative': cumulative,
            'active_steps': active_steps,
            'blended_baseline_cost': round(blended_baseline_cost, 0),
            'location_breakdown': location_breakdown,
            'impl_breakdown': impl_breakdown,
            'cx_revenue_impact': cx_revenue_impact,
            'cost_of_inaction': cost_of_inaction,
            'role_breakdown': role_breakdown,
            'attrition_timeline': attrition_timeline,
        },
        'yearly_data': yearly_data,
        'initiative_impacts': initiative_impacts,
        'pool_snapshots': pool_snapshots,
        'kpi_gaps': kpi_gaps,
        'recommendations': recs,
        'validation': validation,
        'risk_register': risk_register,
        'warnings': [],
    }


# ═══════════════════════════════════════════════════════
# INITIATIVE IMPACTS (uses pools for consistency)
# ═══════════════════════════════════════════════════════
def calc_initiative_impacts(enterprise, all_queues, horizon_months, active_steps,
                            enterprise_fte=0, fte_scale=1.0, fixed_baseline_ftes=None):
    impacts = []; all_inits = []
    blended_cost = _calc_blended_cost(enterprise)
    if 'automation_ai' in active_steps:
        for init in enterprise.get('initiatives_auto', []):
            all_inits.append({**init, 'layer': 'Automation & AI'})
    if 'opmodel' in active_steps:
        for init in enterprise.get('initiatives_opmodel', []):
            all_inits.append({**init, 'layer': 'Operating Model'})

    for init in all_inits:
        vol_impact = fte_impact = aht_impact = cost_impact = 0
        queues_impacted = 0
        for idx, (bu, q) in enumerate(all_queues):
            if not _init_applies_to_queue(init, q): continue
            levers = init.get('levers', [])
            has = False
            if not levers and init.get('lever'):
                levers = [{'lever': init['lever'], 'process_impacts': {'_all': init.get('impact_pct', 0)}}]
            for le in levers:
                if _get_lever_impact_for_queue(le, q) > 0: has = True; break
            if not has: continue
            queues_impacted += 1
            baseline = calc_queue_baseline(q)
            anchored_fte = fixed_baseline_ftes[idx] if fixed_baseline_ftes else baseline['fte']
            # Use simple single-initiative impact (no pool consumption for per-initiative display)
            vq = copy.deepcopy(q)
            dummy_pools = {lever: {'ceiling_fte': 99999, 'consumed_fte': 0, 'remaining_fte': 99999,
                                    'ceiling_contacts': 999999, 'consumed_contacts': 0,
                                    'remaining_contacts': 999999, 'initiatives_consuming': 0}
                          for lever in POOL_TYPES}
            single = apply_initiatives_pooled(vq, [init], horizon_months, dummy_pools, dummy_pools)
            ratio = single['fte'] / baseline['fte'] if baseline['fte'] > 0 else 1.0
            single_fte = anchored_fte * ratio
            vol_impact += baseline['volume'] - single['volume']
            aht_impact += baseline['aht'] - single['aht']
            fte_impact += anchored_fte - single_fte
            cost_impact += (anchored_fte - single_fte) * blended_cost / 12

        lever_names = [le['lever'] for le in init.get('levers', [])] or [init.get('lever', '')]
        impacts.append({
            'name': init.get('name', ''), 'layer': init.get('layer', ''),
            'levers': lever_names, 'lever': ', '.join(lever_names),
            'channels': init.get('eligible_channels', []),
            'queues_impacted': queues_impacted,
            'volume_reduction': round(vol_impact, 0),
            'aht_reduction_mins': round(aht_impact, 2),
            'fte_reduction': round(fte_impact, 2),
            'annual_saving': round(cost_impact * 12, 0),
            'ramp_year1': init.get('ramp_year1', 0.40),
            'ramp_year2': init.get('ramp_year2', 0.80),
            'ramp_year3': init.get('ramp_year3', 1.00),
        })

    loc_moves = enterprise.get('location_strategy', [])
    if loc_moves and 'location' in active_steps:
        total_loc = 0
        for idx, (bu, q) in enumerate(all_queues):
            bl = calc_queue_baseline(q)
            anchored_fte = fixed_baseline_ftes[idx] if fixed_baseline_ftes else bl['fte']
            bl['fte'] = anchored_fte
            loc = apply_location_savings(bl, loc_moves, q, enterprise.get('salary_rates', {}),
                                          blended_cost, enterprise)
            total_loc += loc['location_saving'] * 12
        impacts.append({
            'name': 'Location Strategy (Combined)', 'layer': 'Location Strategy',
            'levers': ['location'], 'lever': 'location', 'channels': ['all'],
            'queues_impacted': len(all_queues), 'volume_reduction': 0,
            'aht_reduction_mins': 0, 'fte_reduction': 0, 'annual_saving': round(total_loc, 0),
            'ramp_year1': 1.0, 'ramp_year2': 1.0, 'ramp_year3': 1.0,
        })

    impacts.sort(key=lambda x: -abs(x['annual_saving']))
    total_saving = sum(i['annual_saving'] for i in impacts)
    for i in impacts:
        i['pct_of_total'] = round(i['annual_saving'] / max(abs(total_saving), 1) * 100, 1)
    return impacts


def calc_kpi_gaps(enterprise):
    gaps = []
    for kpi in enterprise.get('kpis', []):
        if not kpi.get('enabled', True): continue
        current = kpi.get('current_value') or 0
        benchmark = kpi.get('benchmark_value') or 0
        impact = kpi.get('impact', 'decrease')
        if impact == 'decrease':
            gap = current - benchmark; gap_pct = round(gap / max(abs(benchmark), 0.01) * 100, 1)
            is_gap = current > benchmark
        elif impact == 'increase':
            gap = benchmark - current; gap_pct = round(gap / max(abs(benchmark), 0.01) * 100, 1)
            is_gap = current < benchmark
        else:
            gap = abs(current - benchmark); gap_pct = round(gap / max(abs(benchmark), 0.01) * 100, 1)
            is_gap = abs(gap) > benchmark * 0.05
        severity = 'ok'
        if is_gap:
            if abs(gap_pct) > 30: severity = 'critical'
            elif abs(gap_pct) > 15: severity = 'high'
            elif abs(gap_pct) > 5: severity = 'medium'
            else: severity = 'low'
        gaps.append({'kpi_name': kpi.get('name', ''), 'channel': kpi.get('channels', ['all']),
                    'process': kpi.get('processes', ['all']), 'current': current, 'benchmark': benchmark,
                    'impact': impact, 'gap': round(gap, 2), 'gap_pct': gap_pct, 'severity': severity, 'is_gap': is_gap})
    return gaps

def generate_recommendations(kpi_gaps, enterprise):
    existing = {i['name'] for i in enterprise.get('initiatives_auto', [])} | {i['name'] for i in enterprise.get('initiatives_opmodel', [])}
    gap_lever = {'AHT': ['aht_reduction'], 'FCR': ['repeat_reduction'], 'Concurrency': ['concurrency_uplift'],
        'Shrinkage': ['shrinkage_reduction'], 'Occupancy': ['occupancy_uplift'], 'Deflection': ['deflection'],
        'Transfer': ['aht_reduction'], 'Abandon': ['deflection'], 'Repeat': ['repeat_reduction'],
        'ACW': ['acw_reduction'], 'Schedule': ['schedule_efficiency'], 'Containment': ['deflection']}
    recs = []
    for gap in kpi_gaps:
        if not gap['is_gap'] or gap['severity'] in ('ok', 'low'): continue
        matched = []
        for p, lvs in gap_lever.items():
            if p.lower() in gap['kpi_name'].lower(): matched.extend(lvs)
        if not matched: continue
        for init in AUTOMATION_AI_LIBRARY + OPMODEL_LIBRARY:
            if init['name'] in existing: continue
            init_levers = [le['lever'] for le in init.get('levers', [])]
            if not any(lv in matched for lv in init_levers): continue
            imp = max((le.get('process_impacts', {}).get('_all', 0) for le in init.get('levers', [])), default=0)
            rec = {'gap': gap['kpi_name'], 'severity': gap['severity'], 'initiative': init['name'],
                  'lever': ', '.join(init_levers), 'impact': imp, 'complexity': init.get('complexity', 'medium'),
                  'rationale': f"{gap['kpi_name']} gap ({gap['severity']})"}
            if not any(r['initiative'] == rec['initiative'] and r['gap'] == rec['gap'] for r in recs):
                recs.append(rec)
    recs.sort(key=lambda r: ({'critical': 0, 'high': 1, 'medium': 2}.get(r['severity'], 3), -r['impact']))
    return recs[:20]

def run_validation(enterprise, all_queues, yearly_data, active_steps):
    warns = []
    eft = enterprise.get('enterprise_fte', 0)
    tbu = sum(bu.get('current_fte', 0) for bu in enterprise.get('business_units', []))
    if eft > 0 and tbu > 0 and abs(eft - tbu) > 1:
        warns.append(f"Enterprise FTE ({eft}) ≠ BU FTEs ({tbu})")
    if eft <= 0 and tbu <= 0:
        warns.append("No Enterprise FTE defined")
    lm = enterprise.get('location_mix', {})
    tl = sum(v.get('fte', 0) for v in lm.values())
    if tl > 0 and eft > 0 and abs(tl - eft) > 1:
        warns.append(f"Location Mix FTE ({tl}) ≠ Enterprise FTE ({eft})")
    for bu in enterprise.get('business_units', []):
        stated = bu.get('total_monthly_volume', 0)
        actual = sum(q.get('monthly_volume', 0) for q in bu.get('queues', []))
        if stated > 0 and abs(stated - actual) > stated * 0.05:
            warns.append(f"{bu['bu_name']}: Volume gap {abs(stated - actual):,}")
    for bu, q in all_queues:
        qn = q.get('queue_name', '')
        if q.get('chat_concurrency', 1) > 6: warns.append(f"{qn}: Concurrency>6")
        if q.get('occupancy_target', 0) > 0.95: warns.append(f"{qn}: Occupancy>95%")
    for init in enterprise.get('initiatives_auto', []) + enterprise.get('initiatives_opmodel', []):
        mx = 0
        for le in init.get('levers', []):
            mx = max(mx, max(le.get('process_impacts', {}).values(), default=0))
        if not init.get('levers'): mx = init.get('impact_pct', 0)
        if mx > 0.50: warns.append(f"Initiative '{init['name']}': impact {mx * 100:.0f}%>50%")
    if yearly_data:
        y1 = yearly_data[0]
        if y1['baseline_fte'] > 0:
            rp = y1['fte_reduction'] / y1['baseline_fte']
            if rp > 0.60: warns.append(f"Y1 FTE reduction {rp * 100:.0f}%>60%")
            elif rp > 0.40: warns.append(f"Y1 FTE reduction {rp * 100:.0f}%>40%")
    return {'warnings': warns, 'warning_count': len(warns),
            'has_queues': len(all_queues) > 0,
            'has_initiatives': len(enterprise.get('initiatives_auto', [])) + len(enterprise.get('initiatives_opmodel', [])) > 0,
            'has_technology': len(enterprise.get('technology', [])) > 0,
            'has_kpis': len(enterprise.get('kpis', [])) > 0,
            'has_location': len(enterprise.get('location_strategy', [])) > 0,
            'active_steps': active_steps}


# ═══════════════════════════════════════════════════════
# API ROUTES
# ═══════════════════════════════════════════════════════
@app.route('/')
def index(): return send_file(os.path.join(app.root_path, 'templates', 'index.html'))

@app.route('/api/enterprise', methods=['GET'])
def get_ent():
    e = _get_store()
    return jsonify({k: v for k, v in e.items() if k != '_last_result'})

@app.route('/api/enterprise', methods=['PUT'])
def put_ent():
    d = request.json or {}; e = _get_store()
    for k in ['program_name', 'objective', 'waterfall_order', 'waterfall_active', 'planning_horizon_years',
              'discount_rate', 'global_volume_growth_pct', 'global_wage_inflation_pct', 'attrition_rate_monthly',
              'redeployment_pct', 'enterprise_fte', 'currency', 'custom_channels', 'custom_processes',
              'location_mix', 'salary_rates', 'sourcing_mix', 'cx_revenue', 'implementation_costs',
              'scenario_multipliers', 'roles']:
        if k in d: e[k] = d[k]
    return jsonify({'ok': True})

@app.route('/api/bus', methods=['GET'])
def get_bus(): return jsonify(_get_store().get('business_units', []))

@app.route('/api/bus', methods=['POST'])
def add_bu():
    e = _get_store(); d = request.json or {}
    bu = default_bu(d.get('bu_name', f"BU {len(e['business_units']) + 1}"))
    for k in d:
        if k not in ('id', 'queues'): bu[k] = d[k]
    if d.get('volume_matrix'):
        bu['queues'] = []
        for process, channels in d['volume_matrix'].items():
            for channel, volume in channels.items():
                if volume and volume > 0:
                    bu['queues'].append(default_queue(channel, process, volume, bu))
        bu['total_monthly_volume'] = sum(q['monthly_volume'] for q in bu['queues'])
    e['business_units'].append(bu); return jsonify(bu)

@app.route('/api/bus/<bid>', methods=['PUT'])
def put_bu(bid):
    e = _get_store(); d = request.json or {}
    for bu in e['business_units']:
        if bu['id'] == bid:
            for k, v in d.items():
                if k not in ('id',): bu[k] = v
            if 'volume_matrix' in d:
                bu['queues'] = []
                for process, channels in d['volume_matrix'].items():
                    for channel, volume in channels.items():
                        if volume and volume > 0:
                            bu['queues'].append(default_queue(channel, process, volume, bu))
                bu['total_monthly_volume'] = sum(q['monthly_volume'] for q in bu['queues'])
            return jsonify(bu)
    return jsonify({'error': 'Not found'}), 404

@app.route('/api/bus/<bid>', methods=['DELETE'])
def del_bu(bid):
    e = _get_store(); e['business_units'] = [bu for bu in e['business_units'] if bu['id'] != bid]
    return jsonify({'ok': True})

@app.route('/api/bus/<bid>/queues', methods=['GET'])
def get_queues(bid):
    for bu in _get_store().get('business_units', []):
        if bu['id'] == bid: return jsonify(bu.get('queues', []))
    return jsonify([])

@app.route('/api/bus/<bid>/queues/<qid>', methods=['PUT'])
def put_queue(bid, qid):
    d = request.json or {}
    for bu in _get_store().get('business_units', []):
        if bu['id'] == bid:
            for q in bu.get('queues', []):
                if q['id'] == qid:
                    for k, v in d.items():
                        if k not in ('id', 'channel', 'process_tag', 'monthly_volume'):
                            q[k] = v
                    return jsonify(q)
    return jsonify({'error': 'Not found'}), 404

@app.route('/api/queue-processes', methods=['GET'])
def get_queue_processes():
    e = _get_store(); procs = set()
    for bu in e.get('business_units', []):
        for q in bu.get('queues', []): procs.add(q.get('process_tag', ''))
    return jsonify(sorted(list(procs)))

@app.route('/api/process-config', methods=['GET'])
def get_process_config():
    return jsonify({
        'complexity_defaults': PROCESS_COMPLEXITY_DEFAULTS,
        'pool_types': POOL_TYPES,
        'stepped_realization': STEPPED_REALIZATION,
    })

@app.route('/api/roles', methods=['GET'])
def get_roles():
    return jsonify(_get_store().get('roles', ROLE_DEFAULTS))

@app.route('/api/roles', methods=['PUT'])
def put_roles():
    d = request.json or []
    e = _get_store(); e['roles'] = d
    return jsonify({'ok': True})

@app.route('/api/kpis', methods=['GET'])
def get_kpis(): return jsonify(_get_store().get('kpis', []))

@app.route('/api/kpis', methods=['POST'])
def add_kpi():
    e = _get_store(); d = request.json or {}
    kpi = {'id': _uid(), 'name': d.get('name', ''), 'unit': d.get('unit', ''),
           'channels': d.get('channels', ['all']), 'processes': d.get('processes', ['all']),
           'current_value': d.get('current_value', 0), 'benchmark_value': d.get('benchmark_value', 0),
           'impact': d.get('impact', 'decrease'), 'enabled': d.get('enabled', True),
           'category': d.get('category', 'custom')}
    e.setdefault('kpis', []).append(kpi); return jsonify(kpi)

@app.route('/api/kpis/<kid>', methods=['PUT'])
def put_kpi(kid):
    d = request.json or {}
    for kpi in _get_store().get('kpis', []):
        if kpi['id'] == kid:
            for k, v in d.items():
                if k != 'id': kpi[k] = v
            return jsonify(kpi)
    return jsonify({'error': 'Not found'}), 404

@app.route('/api/kpis/<kid>', methods=['DELETE'])
def del_kpi(kid):
    e = _get_store(); e['kpis'] = [k for k in e.get('kpis', []) if k['id'] != kid]
    return jsonify({'ok': True})

@app.route('/api/kpi-library', methods=['GET'])
def get_kpi_lib(): return jsonify(default_kpi_library())

@app.route('/api/kpis/auto-populate', methods=['POST'])
def auto_populate_kpis():
    e = _get_store(); lib = default_kpi_library()
    channels = set(); processes = set(); all_queues = []
    for bu in e.get('business_units', []):
        channels.update(bu.get('channels', [])); processes.update(bu.get('processes', []))
        for q in bu.get('queues', []): all_queues.append(q)
    if not channels: channels = set(CHANNELS)
    if not processes: processes = set(PROCESSES)
    def _qa(field, cf=None, w=False):
        m = [q for q in all_queues if (cf is None or q['channel'] in cf)]
        if not m: return 0
        if w:
            tv = sum(q['monthly_volume'] for q in m)
            return sum(q.get(field, 0) * q['monthly_volume'] for q in m) / tv if tv > 0 else 0
        return sum(q.get(field, 0) for q in m) / len(m)
    cvm = {('AHT', 'voice'): round(_qa('handle_time_minutes', ['voice'], True), 2),
           ('AHT', 'chat'): round(_qa('handle_time_minutes', ['chat'], True), 2),
           ('AHT', 'email'): round(_qa('handle_time_minutes', ['email'], True), 2),
           ('FCR',): round(_qa('fcr_pct', ['voice', 'chat']) * 100, 1),
           ('CSAT',): round(_qa('csat_score'), 1),
           ('Transfer Rate',): round(_qa('transfer_pct', ['voice', 'chat']) * 100, 1),
           ('Abandon Rate',): round(_qa('abandon_rate', ['voice']) * 100, 1),
           ('Chat Concurrency',): round(_qa('chat_concurrency', ['chat']), 1),
           ('SLA %',): round(_qa('sla_target', ['voice', 'chat']) * 100, 1),
           ('ACW',): round(_qa('after_call_work_minutes', ['voice', 'chat'], True), 2),
           ('Voice Occupancy',): round(_qa('occupancy_target', ['voice']) * 100, 1),
           ('Chat Occupancy',): round(_qa('occupancy_target', ['chat']) * 100, 1),
           ('Shrinkage',): round(_qa('shrinkage_pct') * 100, 1),
           ('Schedule Efficiency',): round(_qa('schedule_efficiency') * 100, 1),
           ('Attrition Rate',): round(e.get('attrition_rate_monthly', 0.03) * 100, 1),
           ('Repeat Contact Rate',): round(_qa('repeat_contact_pct', ['voice', 'chat']) * 100, 1)}
    def _gc(n, chs):
        for ch in chs:
            if ch != 'all' and (n, ch) in cvm: return cvm[(n, ch)]
        return cvm.get((n,), 0)
    added = 0
    existing = {(k['name'], str(k.get('channels', [])), str(k.get('processes', []))) for k in e.get('kpis', [])}
    for kd in lib:
        kc = kd.get('channels', ['all']); kp = kd.get('processes', ['all'])
        if 'all' not in kc and not set(kc).intersection(channels): continue
        if 'all' not in kp and not set(kp).intersection(processes): continue
        key = (kd['name'], str(kc), str(kp))
        if key in existing: continue
        kpi = {'id': _uid(), 'name': kd['name'], 'unit': kd.get('unit', ''), 'channels': kc, 'processes': kp,
               'current_value': _gc(kd['name'], kc), 'benchmark_value': kd.get('benchmark', 0),
               'impact': kd.get('impact', 'decrease'), 'enabled': True, 'category': kd.get('category', 'channel')}
        e.setdefault('kpis', []).append(kpi); added += 1
    return jsonify({'added': added, 'total': len(e.get('kpis', []))})

@app.route('/api/channels-processes', methods=['GET'])
def get_channels_processes():
    e = _get_store()
    return jsonify({'channels': CHANNELS + e.get('custom_channels', []),
                    'processes': PROCESSES + e.get('custom_processes', []),
                    'custom_channels': e.get('custom_channels', []),
                    'custom_processes': e.get('custom_processes', [])})

@app.route('/api/channels-processes', methods=['POST'])
def add_channel_process():
    e = _get_store(); d = request.json or {}
    if d.get('type') == 'channel':
        name = d.get('name', '').strip().lower()
        if name and name not in CHANNELS and name not in e.get('custom_channels', []):
            e.setdefault('custom_channels', []).append(name)
            return jsonify({'ok': True, 'added': name})
    elif d.get('type') == 'process':
        name = d.get('name', '').strip().lower().replace(' ', '_')
        if name and name not in PROCESSES and name not in e.get('custom_processes', []):
            e.setdefault('custom_processes', []).append(name)
            return jsonify({'ok': True, 'added': name})
    return jsonify({'ok': False, 'error': 'Already exists or invalid'})

@app.route('/api/initiative-library/<lib_type>', methods=['GET'])
def get_init_lib(lib_type):
    if lib_type == 'auto': return jsonify(AUTOMATION_AI_LIBRARY)
    elif lib_type == 'opmodel': return jsonify(OPMODEL_LIBRARY)
    return jsonify([])

@app.route('/api/initiatives/<init_type>', methods=['GET'])
def get_inits(init_type):
    key = 'initiatives_auto' if init_type == 'auto' else 'initiatives_opmodel'
    return jsonify(_get_store().get(key, []))

@app.route('/api/initiatives/<init_type>', methods=['POST'])
def add_init(init_type):
    e = _get_store(); d = request.json or {}
    key = 'initiatives_auto' if init_type == 'auto' else 'initiatives_opmodel'
    init = {
        'id': _uid(), 'name': d.get('name', ''), 'levers': d.get('levers', []),
        'eligible_channels': d.get('eligible_channels', ['all']),
        'adoption_pct': d.get('adoption_pct', 0.80),
        'start_month': d.get('start_month', 1),
        'ramp_year1': d.get('ramp_year1', 0.40),
        'ramp_year2': d.get('ramp_year2', 0.80),
        'ramp_year3': d.get('ramp_year3', 1.00),
        'complexity': d.get('complexity', 'medium'),
        'wave': d.get('wave', 1),
        'description': d.get('description', ''),
        'risk_category': d.get('risk_category', 'general'),
        'risk_likelihood': d.get('risk_likelihood', 0.3),
        'risk_impact': d.get('risk_impact', 0.3),
    }
    if not init['levers'] and d.get('lever'):
        procs = d.get('eligible_processes', ['all']); imp = d.get('impact_pct', 0)
        pi = {'_all': imp} if ('all' in procs or not procs) else {p: imp for p in procs}
        init['levers'] = [{'lever': d['lever'], 'process_impacts': pi}]
    if d.get('source_channel'): init['source_channel'] = d['source_channel']
    if d.get('target_channel'): init['target_channel'] = d['target_channel']
    e.setdefault(key, []).append(init); return jsonify(init)

@app.route('/api/initiatives/<init_type>/<iid>', methods=['PUT'])
def put_init(init_type, iid):
    d = request.json or {}
    key = 'initiatives_auto' if init_type == 'auto' else 'initiatives_opmodel'
    for init in _get_store().get(key, []):
        if init['id'] == iid:
            for k, v in d.items():
                if k != 'id': init[k] = v
            return jsonify(init)
    return jsonify({'error': 'Not found'}), 404

@app.route('/api/initiatives/<init_type>/<iid>', methods=['DELETE'])
def del_init(init_type, iid):
    e = _get_store(); key = 'initiatives_auto' if init_type == 'auto' else 'initiatives_opmodel'
    e[key] = [i for i in e.get(key, []) if i['id'] != iid]; return jsonify({'ok': True})

@app.route('/api/location-strategy', methods=['GET'])
def get_loc(): return jsonify(_get_store().get('location_strategy', []))

@app.route('/api/location-strategy', methods=['POST'])
def add_loc():
    e = _get_store(); d = request.json or {}
    move = {'id': _uid(), 'from_location': d.get('from_location', 'onshore'),
            'to_location': d.get('to_location', 'nearshore'),
            'processes': d.get('processes', ['all']), 'channels': d.get('channels', ['all']),
            'move_pct': d.get('move_pct', 0.20), 'wave': d.get('wave', 1),
            'start_month': d.get('start_month', 1), 'end_month': d.get('end_month', 12)}
    e.setdefault('location_strategy', []).append(move); return jsonify(move)

@app.route('/api/location-strategy/<mid>', methods=['PUT'])
def put_loc(mid):
    d = request.json or {}
    for m in _get_store().get('location_strategy', []):
        if m['id'] == mid:
            for k, v in d.items():
                if k != 'id': m[k] = v
            return jsonify(m)
    return jsonify({'error': 'Not found'}), 404

@app.route('/api/location-strategy/<mid>', methods=['DELETE'])
def del_loc(mid):
    e = _get_store()
    e['location_strategy'] = [m for m in e.get('location_strategy', []) if m['id'] != mid]
    return jsonify({'ok': True})

@app.route('/api/technology-library', methods=['GET'])
def get_tech_lib(): return jsonify(TECHNOLOGY_LIBRARY)

@app.route('/api/technology', methods=['GET'])
def get_tech(): return jsonify(_get_store().get('technology', []))

@app.route('/api/technology', methods=['POST'])
def add_tech():
    e = _get_store(); d = request.json or {}
    tech = {'id': _uid(), 'name': d.get('name', ''), 'category': d.get('category', 'other'),
            'cost_type': d.get('cost_type', 'both'), 'one_time': d.get('one_time', 0),
            'recurring_monthly': d.get('recurring_monthly', 0), 'start_month': d.get('start_month', 1),
            'end_month': d.get('end_month', 36), 'linked_initiatives': d.get('linked_initiatives', []),
            'description': d.get('description', '')}
    e.setdefault('technology', []).append(tech); return jsonify(tech)

@app.route('/api/technology/<tid>', methods=['PUT'])
def put_tech(tid):
    d = request.json or {}
    for t in _get_store().get('technology', []):
        if t['id'] == tid:
            for k, v in d.items():
                if k != 'id': t[k] = v
            return jsonify(t)
    return jsonify({'error': 'Not found'}), 404

@app.route('/api/technology/<tid>', methods=['DELETE'])
def del_tech(tid):
    e = _get_store()
    e['technology'] = [t for t in e.get('technology', []) if t['id'] != tid]
    return jsonify({'ok': True})

@app.route('/api/run', methods=['POST'])
def run():
    e = _get_store()
    result = run_waterfall(e)
    try:
        result['scenarios'] = _run_scenario_comparison(e, result)
    except: result['scenarios'] = {}
    try:
        result['sensitivity'] = _run_sensitivity(e, result)
    except: result['sensitivity'] = []
    _store['_last_result'] = result
    return jsonify(result)

@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    r = _last_result()
    if not r: return jsonify({'error': 'Run model first'}), 400
    try:
        from openpyxl import Workbook; from openpyxl.styles import Font, PatternFill
    except ImportError: return jsonify({'error': 'openpyxl not installed'}), 500
    wb = Workbook()
    hf = Font(bold=True, color='FFFFFF'); hfill = PatternFill('solid', fgColor='2E2E38')
    def hdr(ws, cols):
        for i, c in enumerate(cols, 1):
            cell = ws.cell(1, i, c); cell.font = hf; cell.fill = hfill

    # Summary
    ws1 = wb.active; ws1.title = 'Summary'; hdr(ws1, ['Metric', 'Value'])
    s = r.get('summary', {})
    for k, v in s.items():
        if k not in ('cumulative', 'active_steps', 'location_breakdown', 'impl_breakdown',
                     'cx_revenue_impact', 'cost_of_inaction', 'role_breakdown', 'attrition_timeline'):
            ws1.append([k.replace('_', ' ').title(), v])

    # Yearly Data
    active = s.get('active_steps', [])
    ws2 = wb.create_sheet('Yearly Data')
    cols = ['Year', 'Baseline FTE', 'Final FTE', 'FTE Reduction']
    for step in active: cols.append(f"{STEP_LABELS.get(step, step)} Saving")
    cols += ['Total Labor', 'Tech Cost', 'Net Saving']; hdr(ws2, cols)
    for y in r.get('yearly_data', []):
        row = [y['year'], y['baseline_fte'], y['final_fte'], y['fte_reduction']]
        for step in active: row.append(y.get('step_savings', {}).get(step, 0))
        row += [y['total_labor_saving'], y['total_tech'], y['net_saving']]; ws2.append(row)

    # Impacts
    ws3 = wb.create_sheet('Impacts')
    hdr(ws3, ['Initiative', 'Layer', 'Levers', 'Queues', 'FTE↓', 'Annual Saving', '%', 'Ramp Y1', 'Ramp Y2', 'Ramp Y3'])
    for i in r.get('initiative_impacts', []):
        ws3.append([i['name'], i['layer'], i.get('lever', ''), i['queues_impacted'],
                   i['fte_reduction'], i['annual_saving'], i['pct_of_total'],
                   i.get('ramp_year1', ''), i.get('ramp_year2', ''), i.get('ramp_year3', '')])

    # Role Breakdown
    rb = s.get('role_breakdown', [])
    if rb:
        ws_roles = wb.create_sheet('Roles')
        hdr(ws_roles, ['Role', 'Baseline FTE', 'Future FTE', 'Reduction', '% Reduction', 'Cost/FTE', 'Annual Saving'])
        for role in rb:
            ws_roles.append([role['role'], role['baseline_fte'], role['future_fte'],
                           role['reduction'], role['pct_reduction'], role['cost_per_fte'], role['annual_saving']])

    # Pool Utilization
    ps = r.get('pool_snapshots', [])
    if ps:
        ws_pools = wb.create_sheet('Pools')
        hdr(ws_pools, ['Year', 'Lever', 'Ceiling FTE', 'Consumed FTE', 'Remaining FTE', 'Utilization %'])
        for snap in ps:
            for lever, pool in snap['pools'].items():
                ws_pools.append([snap['year'], lever, pool['ceiling_fte'], pool['consumed_fte'],
                               pool['remaining_fte'], pool['utilization_pct']])

    # Scenarios
    scenarios = r.get('scenarios', {})
    if scenarios:
        ws4 = wb.create_sheet('Scenarios')
        hdr(ws4, ['Scenario', 'NPV', 'ROI %', 'Payback Year', 'FTE Reduction', 'Net Benefit'])
        for sn in ['conservative', 'base', 'aggressive']:
            sc = scenarios.get(sn, {})
            ws4.append([sn.title(), sc.get('npv', 0), sc.get('roi_pct', 0),
                       sc.get('payback_year', 'N/A'), sc.get('fte_reduction', 0), sc.get('total_net_benefit', 0)])

    # Sensitivity
    sens = r.get('sensitivity', [])
    if sens:
        ws5 = wb.create_sheet('Sensitivity')
        hdr(ws5, ['Variable', 'Low NPV (-20%)', 'Base NPV', 'High NPV (+20%)', 'Swing'])
        for sv in sens: ws5.append([sv['variable'], sv['low_npv'], sv['base_npv'], sv['high_npv'], sv['swing']])

    # Risk Register
    rr = r.get('risk_register', {})
    if rr.get('risks'):
        ws_risk = wb.create_sheet('Risk Register')
        hdr(ws_risk, ['ID', 'Initiative', 'Category', 'Likelihood', 'Impact', 'Score', 'Rating', 'Mitigation'])
        for risk in rr['risks']:
            ws_risk.append([risk['id'], risk['initiative'], risk['category'],
                          risk['likelihood'], risk['impact'], risk['score'], risk['rating'], risk['mitigation']])

    # Implementation costs
    impl = s.get('impl_breakdown', {})
    if impl.get('grand_total', 0) > 0:
        ws6 = wb.create_sheet('Investment')
        hdr(ws6, ['Category', 'Amount'])
        for k in ['technology_total', 'change_management', 'training', 'integration', 'contingency']:
            ws6.append([k.replace('_', ' ').title(), impl.get(k, 0)])
        ws6.append(['Grand Total', impl.get('grand_total', 0)])

    # CX Revenue
    cx = s.get('cx_revenue_impact', {})
    if cx.get('enabled'):
        ws7 = wb.create_sheet('CX Revenue')
        hdr(ws7, ['Year', 'Churn Rate %', 'Customers Retained', 'Revenue Retained'])
        for y in cx.get('yearly', []):
            ws7.append([f"Y{y['year']}", y['churn_rate'], y['retained_customers'], y['revenue_retained']])

    # Cost of Inaction
    coi = s.get('cost_of_inaction', [])
    if coi:
        ws8 = wb.create_sheet('Cost of Inaction')
        hdr(ws8, ['Year', 'FTE Needed', 'Labor Cost', 'Attrition Cost', 'Total'])
        for c in coi:
            ws8.append([f"Y{c['year']}", c['fte_needed'], c['labor_cost'], c['attrition_cost'], c['total_cost']])

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='cc_business_case_v7.xlsx')


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=8081)
